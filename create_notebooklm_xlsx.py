from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Sheet 1: タイトル一覧 ---
ws1 = wb.active
ws1.title = "タイトル一覧"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

headers = ["#", "タイトル", "作成日", "大分類", "中分類"]
col_widths = [5, 75, 12, 28, 32]

for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.auto_filter.ref = "A1:E1"
ws1.freeze_panes = "A2"

data = [
(1,"🎮 Nintendo Museum: Experience Design and Brand Integration Analysis","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(2,"🏠 IKEA Blueprint: Designing the Narrative of Daily Life","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(3,"🏛️ Bilbao Effect: Architecture as a Transformative Experience Machine","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(4,"💎 Jewel Changi: Transforming Transit into a Global Destination","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(5,"⛩️ Translating Kyoto Heritage into the Pinnacle of Luxury Design","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(6,"🍄 Structural Principles of Immersive Experience Design","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(7,"🌋 The Architectural Narrative and Experience Design of Tokyo DisneySea","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(8,"🏟️ Es Con Field: Redefining Ballpark Experience and Spatial Design Analysis","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(9,"☕ The Starbucks Reserve Roastery: Architecture of Immersive Experience Design","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(10,"🌀 The Architecture of Wandering: Analyzing teamLab Borderless Experience Design","2026/03/30","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(11,"🎮 Nintendo Museum: A Masterclass in Brand Experience Design","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(12,"🗼 Engineering the Ascent: Sky Tree's Vertical Emotional Design","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(13,"🏢 The S-Grand Genealogy: Real Estate Innovation and Activist Capital","2026/03/29","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(14,"🌳 マールブランシュ ロマンの森：商業建築における物語性と空間設計の考察","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(15,"🏢 The S-Grant Legacy: Evolution of Japanese Real Estate and Capital","2026/03/29","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(16,"🗼 The Vertical Narrative: Engineering the Tokyo Skytree Experience","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(17,"🏟️ Escon Field Hokkaido: The Evolution of Modern Ballpark Design","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(18,"🏟️ Nagasaki Stadium City: A Paradigm Shift in Urban Development","2026/03/29","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(19,"🎂 The Edible Monument: The Artistic Evolution of Cake","2026/03/29","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(20,"📖 The Geometry of Travel: Visual Design and Reader Navigation","2026/03/29","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(21,"📖 The Architecture of Visual Journeys in Modern Travel Media","2026/03/29","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(22,"🌱 The Architecture of Benevolence: Taneya's Visionary Retail Philosophy","2026/03/29","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(23,"🖼️ Framing the Feminine: Evolution of Composition and Psychological Representation","2026/03/29","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(24,"👁️ The Mossad Doctrine: Strategic Intelligence and Special Operations","2026/03/29","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(25,"🌅 The Rising Phoenix: Hiroyuki Sugimoto and the Rebirth of Syla Technologies","2026/03/29","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(26,"🏢 The Evolution and Legacy of the S-Grand Business Model","2026/03/29","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(27,"🖼️ The Architecture of the Gaze: Composition in Portraiture","2026/03/29","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(28,"🚀 Gemini and Google Workspace: The 2026 Agentic Workflow Revolution","2026/03/28","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(29,"🧋 Gong cha: The Strategy and Future of a Tea Empire","2026/03/28","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(30,"🏗️ Applied Materials: The Atomic Architect of the AI Era","2026/03/28","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(31,"📈 Intuit: The AI-Driven Evolution of Financial Ecosystems","2026/03/28","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(32,"🍱 かつや：食のイノベーションと市場支配の数理分析","2026/03/28","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(33,"🔪 Forging Precision: The Science and Heritage of Japanese Blades","2026/03/28","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(34,"🐷 かつや：圧倒的シェアを築く破壊的イノベーションの全貌","2026/03/28","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(35,"🛍️ Shopify and the Global Democratization of Commerce","2026/03/28","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(36,"🏗️ Applied Materials: The Strategic Mastery of Materials Engineering","2026/03/28","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(37,"🦍 The Engineering and Mystery of Go Go Curry","2026/03/28","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(38,"🍶 Japanese Shoyu: The Art and Science of Premium Brewing","2026/03/28","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(39,"🔪 Japanese Culinary Metallurgy: The Science of Precision Cutlery","2026/03/28","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(40,"🍶 The Evolution and Biotechnology of Premium Soy Sauce Manufacturing","2026/03/28","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(41,"🍛 The Engineering and Gastronomy of Kanazawa Curry","2026/03/28","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(42,"💩 Curried Submission: The Documentary Portrait","2026/03/27","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(43,"🏗️ Tokyo Timber and Tradition: A Craftsman's 2026 Vision Guide","2026/03/27","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(44,"💪 RIZAP: The Mechanics of Behavioral Change and Result Commitment","2026/03/27","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(45,"🌌 放射冷却の物理的原理と多角的応用技術の全貌","2026/03/27","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(46,"🌡️ Japan's Warming Seas: The Great Fisheries Shift","2026/03/27","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(47,"✍️ IDE Ecosystems for Modern Novel Writing","2026/03/26","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(48,"👁️ チラシの作り方","2026/03/24","2. ビジネス・経営戦略","2-5. 広告・マーケティング"),
(49,"✍️ Architecting the Modern Novel: Integrated Workflows for Digital Authorship","2026/03/26","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(50,"🍣 The Sea's Legacy: Japan's Evolutionary Path to Modern Meat Culture","2026/03/26","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(51,"👁️ Architects of Intelligence: Data Ontology and Industrial Disruption","2026/03/26","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(52,"🎬 Claude Code: Autonomous Video Editing Architecture and Implementation Guide","2026/03/26","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(53,"🎵 Suno AI: The Architecture and Evolution of Audio Synthesis","2026/03/26","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(54,"📈 Kyoto On-Demand Platform Demand and Supply Dynamics","2026/03/26","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(55,"🐟 The Anatomy of Japanese Fish Culture","2026/03/26","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(56,"🚢 The Engineering Evolution of the Asuka Luxury Cruise Fleet","2026/03/26","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(57,"🌎 The Blue Tide: Geopolitical Realignment in Latin America","2026/03/26","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(58,"🚢 The Structural Evolution of the Asuka Cruise Ship Series","2026/03/26","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(59,"👁️ Data Ontology and the Logic of Industrial Disruption","2026/03/26","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(60,"✍️ A Shadow Partner: The 25th Anniversary Script Concept","2026/03/26","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(61,"🎬 Shaft at 50: The Evolution of an Animation Aesthetic","2026/03/26","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(62,"🍚 The Pathology of Carbohydrate Fatigue and Locabo Intervention Strategies","2026/03/25","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(63,"⚾️ Driveline Mechanics: The Data-Driven Evolution of Baseball Performance","2026/03/25","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(64,"❄️ Cold Chain Evolution: From Thermodynamics to the Final Frontier","2026/03/25","4. 飲食・食品産業","4-3. 冷凍・冷蔵・物流"),
(65,"🍿 Netflix: The Data-Driven Art of the Teaser Strategy","2026/03/25","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(66,"👤 Kagerou Kagerou: The Shadow Cicada Legend","2026/03/25","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(67,"🍚 牛丼チェーン","2026/03/24","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(68,"🚀 Frontier Engineering: The Evolution of Forward Deployed Roles","2026/03/24","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(69,"📈 Goldman Sachs 3.0: The AI-Driven Financial Evolution","2026/03/24","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(70,"🍴 飲食チェーン多業態化の構造障壁と成功の戦略的条件","2026/03/23","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(71,"🐼 The Panda Way: Engineering the American Chinese Cuisine Empire","2026/03/23","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(72,"📣 広告代理店の本質的役割と現代的変容：DX時代の戦略パートナー","2026/03/23","2. ビジネス・経営戦略","2-5. 広告・マーケティング"),
(73,"🖼️ The Modern Curator: Bridging Cultural Heritage and Social Mission","2026/03/23","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(74,"📈 The Science of Continuity: A Strategic Guide to Habit Formation","2026/03/23","8. 教育・学習","8-2. 教育理論・AI教育"),
(75,"🍜 Gift of the Gavel: Scaling the Yokohama Ramen Empire","2026/03/23","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(76,"🚢 Choke Points and the New Era of Maritime Geopolitics","2026/03/23","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(77,"🛢️ The Hydrocarbon Dichotomy: A Comparative Study of Crude and LNG","2026/03/23","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(78,"🕊️ The Genesis and Evolution of the Liberal Arts","2026/03/23","8. 教育・学習","8-2. 教育理論・AI教育"),
(79,"🌇 Analysis of Temporal Traffic Flow and Accident Correlation in Japan","2026/03/23","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(80,"⛽ Japan Gas Station Construction Standards and Investment Costs","2026/03/23","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(81,"🍜 Nissin Foods: Evolution of Instant Ramen and Global Food Innovation","2026/03/23","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(82,"⛽ Japan Gas Station Construction Standards and Investment Costs (2)","2026/03/23","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(83,"🏪 Japanese Convenience Store Construction and Market Outlook 2025-2026","2026/03/23","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(84,"🍱 The Multi-Brand Dilemma: Structural Barriers in Restaurant Chain Expansion","2026/03/23","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(85,"🦀 Bakemonogatari: The Psychology of Monsters and Adolescence","2026/03/22","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(86,"👁️ パランティア・テクノロジーズ","2026/03/22","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(87,"🚀 2026年における複数のAIエージェントを組み合わせた戦略的運用フレームワーク","2026/03/22","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(88,"🔋 ARMアーキテクチャが他のプロセッサと比較して圧倒的な省電力性を実現している技術的背景","2026/03/22","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(89,"🔋 ARMアーキテクチャがなぜ他の設計に比べて圧倒的な省電力性を実現できるのか","2026/03/22","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(90,"🏗️ パランティアと識学","2026/03/22","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(91,"🏠 Dual-Region Living: The Reality of Two Homes","2026/03/22","10. ライフスタイル・地方創生","10-1. 二拠点生活・地方創生"),
(92,"🏠 The Reality of Dual-Regional Living","2026/03/22","10. ライフスタイル・地方創生","10-1. 二拠点生活・地方創生"),
(93,"🏗️ Structural Realism: Synchronizing Shichigaku Consciousness and Palantir Ontology","2026/03/22","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(94,"🚀 Strategic Multi-AI Orchestration Framework 2026","2026/03/21","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(95,"💥 The Geopolitical Kinetic Shift and the Collapse of CRINK Strategy","2026/03/03","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(96,"🔗 Strategic Integration of Value and Supply Chains","2026/03/21","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(97,"🏗️ パランティア","2026/03/20","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(98,"🛠️ The Bit and Atom Revolution: Remaking the Global Industrial Order","2026/03/20","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(99,"🏗️ AI Strategy for Architectural Design and Construction Management","2026/03/20","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(100,"🎬 脚本原作監督エンタメ","2026/03/21","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(101,"🗾 The G/L Theory and Japan's AI Transformation Strategy","2026/03/20","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(102,"🏗️ 工務店の業務を「読む・考える・書く・確かめる」","2026/03/20","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(103,"🏠 二地域居住：ライフスタイルの変容と地方創生の新たな法的枠組み","2026/03/20","10. ライフスタイル・地方創生","10-1. 二拠点生活・地方創生"),
(104,"🌐 The G and L Worlds: A Blueprint for Japanese Renewal","2026/03/21","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(105,"🗾 飯田と京都：二拠点生活の構造分析と持続可能な居住モデル","2026/03/20","10. ライフスタイル・地方創生","10-1. 二拠点生活・地方創生"),
(106,"🏘️ 飯田と京都を結ぶ二拠点生活の構造と戦略的展望","2026/03/20","10. ライフスタイル・地方創生","10-1. 二拠点生活・地方創生"),
(107,"👁️ Palantir Technologies: The Digital Nerve System of Western Decision Infrastructure","2026/03/20","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(108,"🤖 AIとイラン戦争","2026/03/05","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(109,"👁️ Palantir: Bridging Global Data and Tactical AI Implementation","2026/03/20","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(110,"🦑 イカゲームの全ゲーム","2026/03/09","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(111,"🇰🇷 トリリンガルのトミ 韓国語","2025/09/11","8. 教育・学習","8-1. 語学学習"),
(112,"🛡️ U.S. Defense Technology Strategy: 2026 Fiscal and Strategic Analysis","2026/03/04","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(113,"📉 The Petrodollar Twilight and the Multipolar Financial Transition","2026/03/09","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(114,"🧠 将棋、チェス、囲碁","2026/03/10","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(115,"🦑 The Squid Game Trilogy: A Complete Ritual Analysis","2026/03/09","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(116,"⚖️ Kao vs P&G: Contrast in Consumer Goods Strategy","2026/03/09","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(117,"🗺️ 陣取りゲーム","2026/03/09","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(118,"🚀 Google Antigravity","2026/03/09","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(119,"🧬 SENOX: The Natural Breakthrough in Senolytic Science","2026/03/08","7. 健康・医療・ウェルネス","7-2. 老化・セノリティクス"),
(120,"📚 本出版の工程表（構造問題、ネットに勝てない）","2026/03/01","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(121,"📔 電子レンジ","2026/03/06","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(122,"🚩 チェスと将棋","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(123,"🧩 テトリスPuzzle Games","2026/03/09","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(124,"🗺️ 陣取りゲームConquest Games","2026/03/09","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(125,"🧩 パズルゲーム","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(126,"📔 イランが国家の存続を懸けて構築した地下軍事インフラの構造","2026/03/06","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(127,"⚪ AlphaGo: The Evolution from Game AI to Scientific Discovery","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(128,"📊 Statistical Dynamics of Competitive Balance in Professional Sports","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(129,"🤖 米国によるイラン攻撃の裏側と最新兵器がもたらす地政学的インパクト","2026/03/04","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(130,"📊 The Mechanics of Winning: Competitive Balance Across Professional Sports","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(131,"📔 陣取りゲーム","2026/03/08","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(132,"🌋 2026年の中東情勢","2026/03/04","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(133,"🇮🇱 The 501(c) Strategy: Engineering the 2026 Iran War","2026/03/05","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(134,"🌍 Middle Eastern Geopolitics: From Islamic Origins to the 2026 Crisis","2026/03/03","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(135,"💵 ドル本位性","2026/03/04","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(136,"👁️ Palantir: Architecting the Sovereign AI Operating System","2026/03/04","1. テクノロジー・AI","1-2. Palantir・データ基盤"),
(137,"🤖 The Architecture of Intelligentized Warfare","2026/03/04","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(138,"🌍 Crossroads of Fortune: Middle Eastern Geopolitics and Resource Distribution","2026/03/03","5. 地政学・国際関係","5-1. 中東・軍事・インテリジェンス"),
(139,"🔄 Human Proxies: The Rise of Agentic AI Supremacy","2026/03/01","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(140,"🏗️ AI Management: Leading the Agentic Revolution","2026/03/01","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(141,"🥩 The Physiology of Sustainable Fat Loss and Insulin Optimization","2026/03/01","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(142,"🤖 Claude Code: The Autonomous Paradigm of Agentic Engineering","2026/03/01","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(143,"💻 The Architecture of Modern Computing: Integrating CPU and GPU Paradigms","2026/03/01","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(144,"🥩 The New Science of Sugar Fatigue and Healthy Lipids","2026/03/01","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(145,"💼 Claude and Anthropic: The Business of Practical AI","2026/02/20","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(146,"🍱 Modern Strategies in Japan's 2026 Bento Market Analysis","2026/02/25","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(147,"🏪 The Strategic Architecture of Japan's Convenience Store Giants","2026/02/25","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(148,"💧 The Lifeblood of Kyoto: The Lake Biwa Canal Project","2026/02/25","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(149,"🍊 The Hermès Blueprint: Artisan Excellence and Controlled Rarity","2026/02/24","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(150,"🍜 Marugame Seimen: Digital Transformation and Psychological Capital Strategy","2026/02/24","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(151,"🤖 Claude Cowork: The Evolution of Autonomous AI Agents","2026/02/23","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(152,"🍔 Golden Arches Rising: McDonald's Japan Strategic Growth Analysis","2026/02/23","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(153,"💎 The Structural Essence of Keyence Value Creation","2026/02/23","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(154,"🏭 BMW iFACTORY: The Future of Virtualized Automotive Manufacturing","2026/02/23","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(155,"💊 The Strategic Profit Models of Japanese Drugstores","2026/02/23","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(156,"📦 Unboxing UX: The Strategic Art of Post-Pandemic Delivery Experience","2026/02/19","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(157,"🗺️ The Claude Enterprise Roadmap: Integrating Knowledge and AI Governance","2026/02/19","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(158,"📦 The Architecture of Unboxing: Engineering the Ritual of First Contact","2026/02/19","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(159,"📦 The Amazon Unboxing Experience: Design Philosophy and Packaging Innovation","2026/02/19","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(160,"📦 The Architecture of Anticipation: Decoding Apple's Unboxing UX","2026/02/19","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(161,"🕸️ ドン・フェイラ","2026/02/07","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(162,"🤖 Claude Code Mastery Roadmap: From Foundations to Agentic Development","2026/02/16","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(163,"📈 DIPS: The Engineering of Scientific Management and Sales Power","2026/02/15","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(164,"⚙️ DIPS Theory and the Engineering of Scientific Management","2026/02/15","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(165,"📈 The DIPS Theory and Sales Power Multiplication Program","2026/02/15","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(166,"📈 DIPS: The Science of White-Collar Productivity and Digital Transformation","2026/02/14","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(167,"📦 The Logic of Amazon Packaging and Logistical Algorithms","2026/02/14","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(168,"🧪 The Molecular Dialectics of Water and Oil","2026/02/14","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(169,"🤖 2026 AI Benchmark: OpenAI Google and Anthropic Analysis","2026/02/14","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(170,"🎮 Genesis of the Eight-Bit Paradigm","2026/02/12","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(171,"📊 Claude and the Paradigm of Automated Excel Engineering","2026/02/12","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(172,"🏢 Shaid and the High-Value Strategy of Sha Maison","2026/02/13","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(173,"🏢 The Shinoken Blueprint: Democratizing Real Estate Investment Through Vertical Integration","2026/02/13","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(174,"🏗️ ★Claude Opus 4.6という最新AI技術を活用し日本の地場工務店が抱える業務上の課題","2026/02/13","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(175,"📊 ★「Claude」を活用した次世代のエクセルテンプレート作成と業務自動化","2026/02/12","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(176,"🏆 Anthropic社のClaude 4シリーズの優位性に焦点","2026/02/11","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(177,"🤖 Claude Opus 4.6の登場がソフトウェア業界にもたらす劇的な構造変化","2026/02/13","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(178,"🏗️ Claude Opus 4.6 2026年の日本の建設業界が直面する深刻な労働力不足や法規制の強化","2026/02/13","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(179,"🤖 Tesla 2026: The Physical AI and Robotics Revolution","2026/02/12","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(180,"🚗 Evolution and Economics of Modern Automotive Powertrains","2026/02/12","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(181,"🚀 Muskonomy: The Architect of 21st Century Civilization","2026/02/12","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(182,"🧱 The LEGO System: Engineering Global Play Foundations","2026/02/12","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(183,"👻 The Neutrino Paradigm: Standard Model Extensions and Cosmic Origins","2026/02/12","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(184,"🌌 Evolution of the Kamiokande Experiments: Neutrinos and the Cosmic Frontier","2026/02/12","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(185,"⛏️ The Gold Rush Structure: Infrastructure Economics and Market Dynamics","2026/02/12","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(186,"🚦 Precision Engineering of Modern Urban Traffic Signal Infrastructure","2026/02/12","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(187,"🕹️ The Algorithmic Architecture of Pac-Man","2026/02/12","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(188,"🕹️ Tetris: Computational Architecture and Pedagogical Implementation","2026/02/12","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(189,"🎮 Genesis of the Eight-Bit Paradigm (2)","2026/02/12","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(190,"💗 神田昌典氏の著書『あなたの会社が90日で儲かる！』","2026/02/09","2. ビジネス・経営戦略","2-5. 広告・マーケティング"),
(191,"☀️ アスファルト、コンクリート、土壌","2026/02/09","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(192,"📈 インフレーションのメカニズム","2026/02/07","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(193,"🎬 中島信也：喜んでもらイズムと広告演出の四半世紀","2026/02/09","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(194,"🧠 ジョン・J・レイティ博士の理論に基づき身体活動が脳の構造を物理的に進化させるメカニズム","2026/02/08","7. 健康・医療・ウェルネス","7-3. 運動・脳科学"),
(195,"🦑 Squid Game: The Architecture of Visual Narrative","2026/02/11","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(196,"🎧 YouTubeで高度英語習得モデル","2026/02/05","8. 教育・学習","8-1. 語学学習"),
(197,"🍱 コンビニ弁当とほっともっとに代表される店内調理弁当","2026/02/06","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(198,"🧀 スペンサー・ジョンソン博士の世界的ベストセラー『チーズはどこへ消えた？』","2026/02/07","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(199,"☃️ 『スノーボール』ウォーレン・バフェットの半生と哲学","2026/02/07","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(200,"🏠 2025年の建築基準法改正や資材高騰人手不足といった「四重苦」","2026/02/07","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(201,"📈 神田昌典氏の著書『あなたの会社が90日で儲かる！』","2026/02/09","2. ビジネス・経営戦略","2-5. 広告・マーケティング"),
(202,"🏗️ 地場工務店が木造3階建て共同住宅市場へ参入するための成長戦略","2026/02/06","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(203,"🤖 モラベックのパラドックス","2026/02/08","7. 健康・医療・ウェルネス","7-3. 運動・脳科学"),
(204,"🎨 水野学氏と彼が率いるgood design company","2026/02/07","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(205,"🏦 Too Big to Fail Paradox","2026/02/07","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(206,"🧠 ジョン・J・レイティ博士の研究に基づく身体活動が脳の物理的構造と機能を劇的に変化させる","2026/02/08","7. 健康・医療・ウェルネス","7-3. 運動・脳科学"),
(207,"🥑 リアン・ヴォーゲルによるケトジェニック・ダイエット","2026/02/08","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(208,"📺 佐々木宏の足跡と制作哲学","2026/02/09","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(209,"🥩 北里大学の山田悟氏が提唱する糖質制限","2026/02/08","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(210,"🔋 現代社会のエネルギーインフラを支えるリチウムイオン電池","2026/02/08","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(211,"❄️ リチウムイオン電池が低温下で性能低下を起こす物理化学的なメカニズム","2026/02/09","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(212,"❄️ 冬季にエアコンの電気代が跳ね上がる要因","2026/02/09","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(213,"🎥 澤本嘉光：広告表現の変革とナラティブの構築術","2026/02/09","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(214,"🌀 知性の物理2.0：AIと人間が溶け合う社会の相転移","2026/02/08","7. 健康・医療・ウェルネス","7-3. 運動・脳科学"),
(215,"🎨 水野学氏くまもん","2026/02/07","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(216,"🎨 佐野研二郎にゃんまげLISMO","2026/02/07","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(217,"🏗️ 2025年4月に施行される建築基準法改正","2026/02/06","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(218,"🟥 佐藤可士和：アイコニック・ブランディングと整理術の全貌","2026/02/07","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(219,"🧶 森本千絵ミスチル、ユーミン","2026/02/07","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(220,"❄️ 現代の空調システムにおけるヒートポンプ技術の熱力学的原理","2026/02/02","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(221,"🏄‍♂️ Vibe Coding: The Zero-to-One AI App Development Roadmap","2026/02/05","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(222,"🎧 Digital Scaffolding for Advanced YouTube Language Immersion","2026/02/05","8. 教育・学習","8-1. 語学学習"),
(223,"🏄‍♂️ The Vibe Coding Roadmap: Mastering AI Software Development","2026/02/05","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(224,"📈 Tipping Point","2026/02/05","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(225,"⏳ ティモシー・フェリスの提唱した「週4時間労働」","2026/02/05","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(226,"🏪 日本のコンビニエンスストア業界","2026/02/03","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(227,"🍣 日本の回転寿司業界","2026/02/04","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(228,"🧊 冷蔵庫","2026/02/03","4. 飲食・食品産業","4-3. 冷凍・冷蔵・物流"),
(229,"🍱 日本の持ち帰り弁当業界","2026/02/03","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(230,"📈 顧客生涯価値（LTV）","2026/02/03","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(231,"🍔 日本マクドナルドの経営を担った原田泳幸氏とサラ・カサノバ氏","2026/02/03","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(232,"🚀 イーロン・マスク","2026/02/05","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(233,"🥩 カロリー制限の常識を覆す新しい食事法","2026/02/04","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(234,"🔥 住宅用給湯システム","2026/02/03","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(235,"🥣 牛丼業界における主要三社","2026/02/04","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(236,"🛡️ ワークマンの看板ブランド「イージス」","2026/02/04","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(237,"🦕 ロングテール戦略","2026/02/03","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(238,"🌳 コアコンピタンス経営","2026/02/04","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(239,"🔗 マイケル・ポーターが提唱したバリューチェーン（価値連鎖）","2026/02/04","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(240,"☁️ SaaS、PaaS、IaaSの三層構造","2026/02/04","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(241,"🚀 The Zero to One Philosophy of Vertical Progress","2026/02/05","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(242,"🎨 一人の人間が指揮を執るソロ・スタジオ形態","2026/02/04","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(243,"🪚 木工加工において不可欠なトリマー","2026/02/04","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(244,"📈 The Mechanics and Strategic Dynamics of Social Tipping Points","2026/02/05","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(245,"💹 円キャリートレード","2026/02/04","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(246,"🛠️ 「メイカーズ」","2026/02/04","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(247,"🧵 東レ・ユニクロ・ワークマン：素材革新が拓く新産業構造分析","2026/02/04","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(248,"📜 英語史：綴りと発音の謎を解き明かす学問","2026/02/04","8. 教育・学習","8-1. 語学学習"),
(249,"🌍 AI時代になぜ英語を学ぶのか：言語と文化の認知学","2026/02/04","8. 教育・学習","8-1. 語学学習"),
(250,"📺 OLED vs. Mini LED: The Next Generation Display Analysis","2026/02/02","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(251,"🥗 外食で糖質制限を実現する","2026/02/03","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(252,"⚱️ Ceramics Evolution: From Ancient Pottery to Advanced Materials","2026/02/02","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(253,"🚀 AI-Driven Development: 2025 Frameworks and Strategic Workflows","2026/02/02","1. テクノロジー・AI","1-4. ソフトウェア開発・Vibe Coding"),
(254,"⚱️ Inorganic Material Science: From Traditional Pottery to Modern Ceramics","2026/02/02","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(255,"🌡️ Thermodynamic Analysis of Advanced Heat Pump Systems","2026/02/02","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(256,"☕️ The Unimat Strategy: Evolution of Hospitality and Social Services","2026/02/02","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(257,"💪 The RIZAP Architecture: Science of Metabolic Transformation","2026/02/02","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(258,"🕳️ Crossing the Chasm: The Deep Logic of Tech Diffusion","2026/02/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(259,"📱 Architectural Evolution of Next-Generation Smartphone Systems","2026/02/02","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(260,"❄️ Thermodynamic Principles and Control of Advanced Heat Pump Systems","2026/02/02","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(261,"⚱️ Inorganic Material Science: From Traditional Pottery to Modern Ceramics (2)","2026/02/02","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(262,"🛵 Scooter Engineering and the Evolution of Urban Mobility","2026/02/02","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(263,"📱 Next-Generation Smartphone Architecture and Component Integration Report","2026/02/02","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(264,"🎮 Evolution of the Global Video Game Console Market","2026/02/01","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(265,"🕳️ Crossing the Chasm: Strategies for High-Tech Market Diffusion","2026/02/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(266,"🎮 Evolution of the Console: A History of Video Gaming","2026/02/01","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(267,"🔌 電気工事、電子工作、半導体","2026/02/01","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(268,"🥩 糖質制限ダイエット","2026/02/01","7. 健康・医療・ウェルネス","7-1. 糖質制限・ダイエット"),
(269,"❄️ 冷凍技術","2026/02/01","4. 飲食・食品産業","4-3. 冷凍・冷蔵・物流"),
(270,"🍱 Modern Teishoku Management and Strategic Menu Design","2026/02/01","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(271,"⚡ Powering the AI Revolution: Global Grids and Economic Impact","2026/02/01","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(272,"🍱 Strategic Outlook of the Japanese Bento Industry 2025-2026","2026/02/01","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(273,"🍣 The Evolution of the Japanese Conveyor Belt Sushi Market","2026/02/01","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(274,"📔 Microwave Engineering and Dielectric Heating Analysis","2026/02/01","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(275,"🛠️ ハードウェアハッキング","2026/01/31","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(276,"🏢 シノケン","2026/02/01","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(277,"⚖️ 田口メソッド","2026/02/01","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(278,"🤖 自律型AIエージェント 専門的な部下","2026/02/01","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(279,"🏪 Convenience Store Infrastructure: Decarbonization and Advanced Energy Management","2026/02/01","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(280,"🎨 国宝と重要文化財の違い","2026/01/10","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(281,"📔 Untitled notebook","2026/01/31","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(282,"📈 The LTV Compass: Strategic Frameworks for Long-Term Customer Value","2026/01/31","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(283,"✨ Luxury Redefined: Transforming Offices with Nitori Product Mixing","2026/01/31","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(284,"💹 Hikari Tsushin: From Sales Powerhouse to Capital Alchemist","2026/01/31","2. ビジネス・経営戦略","2-2. 企業分析（テック・金融）"),
(285,"📺 The Sony-TCL Alliance and the Rise of Chinese Tech Dominance","2026/01/31","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(286,"📔 Untitled notebook","2026/01/29","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(287,"🍑 Socioeconomic Infiltration: The Parasite Narrative","2026/01/28","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(288,"🎓 学校では教えてくれない英文法","2025/09/06","8. 教育・学習","8-1. 語学学習"),
(289,"⚡ Japan's Power Grid: From High-Voltage Generation to Smart Distribution","2026/01/28","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(290,"🎾 The Winning Strategy for the Loser's Game","2026/01/28","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(291,"📈 A Random Walk Through Investment Theory and Market Strategy","2026/01/28","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(292,"🏗️ Building Better: The Architect's Guide to In-House Construction","2026/01/28","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(293,"🥘 Chen Kenichi's Home-Style Mapo Tofu Recipe","2026/01/27","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(294,"📈 微分","2026/01/28","8. 教育・学習","8-3. 数学・理系基礎"),
(295,"🚢 コンテナ物語","2026/01/28","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(296,"🎓 英語の先生","2026/01/28","8. 教育・学習","8-1. 語学学習"),
(297,"⛏️ レアメタルレアアース","2026/01/26","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(298,"🏗️ Gemini: Strategic AI Integration in Modern Architectural Renovation","2026/01/27","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(299,"🏗️ Gemini: Strategic AI Integration in Modern Architectural Renovation (2)","2026/01/27","3. 建築・空間デザイン","3-3. 工務店×AI活用"),
(300,"🧠 Intellectual Topographies: Comparing Go Shogi and Chess","2026/01/27","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(301,"📟 半導体戦争","2026/01/01","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(302,"🌍 Davos 2026: The End of the Western World Order","2026/01/26","5. 地政学・国際関係","5-2. グローバル経済・地政学"),
(303,"🏗️ Stroog: Engineering the Future of Urban Timber Architecture","2026/01/15","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(304,"🪵 Modern Timber Revolution: Structural Innovation and Urban Sustainability","2026/01/14","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(305,"📈 マクロ経済学","2026/01/01","8. 教育・学習","8-3. 数学・理系基礎"),
(306,"🏘️ Roadside Leaseback Strategies and Landowner Asset Optimization","2026/01/17","2. ビジネス・経営戦略","2-4. 不動産・投資"),
(307,"💎 The McKinsey Mind: Frameworks for Strategic Problem Solving","2026/01/16","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(308,"🛠️ The 2026 AI Stack: A Blueprint for Professional Survival","2026/01/16","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(309,"📈 The Yotsuya Gakuin Double Education System Analysis Report","2026/01/15","8. 教育・学習","8-2. 教育理論・AI教育"),
(310,"🌿 Tokyu Fudosan: Greening Shibuya for Urban Biodiversity","2026/01/15","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(311,"🏢 Japan's Three-Story Wood Apartment Market and 2025 Regulatory Shift","2026/01/15","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(312,"🏗️ Universal Theory of Design: From Architecture to Semiconductors","2026/01/15","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(313,"📘 君の名は絵コンテ","2026/01/15","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(314,"⏱️ イカゲームパラサイトのリズム","2026/01/15","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(315,"🚉 Japan's Major Railway Terminals: Connectivity and 2026 Urban Evolution","2026/01/15","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(316,"🌀 AI駆動型階層化学習：NotebookLMによる教育変革の理論と実装","2026/01/15","8. 教育・学習","8-2. 教育理論・AI教育"),
(317,"🛠️ Makers: The New Industrial Revolution","2026/01/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(318,"☕ Starbucks Japan: Architecture and Regional Integration Design Report","2026/01/14","3. 建築・空間デザイン","3-1. 体験設計・イマーシブデザイン"),
(319,"⛩️ 京都検定","2026/01/01","8. 教育・学習","8-1. 語学学習"),
(320,"🔋 Scooter Battery Dynamics and Starting Recovery Mechanisms","2026/01/14","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(321,"⚛️ Matter Unveiled: From Atomic Structure to Particle Physics","2026/01/14","6. 科学・工学・エネルギー","6-1. 物理・化学・素材"),
(322,"🚉 Japan's Major Railway Hubs: 2024 Analysis and 2026 Redevelopment Vision","2026/01/14","3. 建築・空間デザイン","3-4. 都市開発・インフラ"),
(323,"⚡ Power Infrastructure: The Evolution of AC and DC Systems","2026/01/14","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(324,"🚀 The Dream Factory: HILLTOP's Revolution in Creative Manufacturing","2026/01/02","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(325,"🏗️ 許容応力度計算の理論体系と実務的適用：2025年法改正への展望","2026/01/05","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(326,"📈 The Universal Law of Learning: From Psychology to AI","2026/01/13","8. 教育・学習","8-2. 教育理論・AI教育"),
(327,"📈 The Challenge to Management: Principles of Decisive Leadership","2026/01/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(328,"🎓 AI-Enhanced Pedagogy: Transforming Learning with Gemini and NotebookLM","2026/01/14","8. 教育・学習","8-2. 教育理論・AI教育"),
(329,"📱 スマホの仕組み","2026/01/13","1. テクノロジー・AI","1-5. デバイス・ディスプレイ"),
(330,"🔓 Retrofitting LIXIL Doors with Aiphone Intercom Interoperability Systems","2026/01/13","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(331,"⛩️ 建築工法","2026/01/05","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(332,"🔋 リチウムイオン電池の技術的進化と次世代蓄電技術の展望","2026/01/12","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(333,"📈 数学","2025/12/04","8. 教育・学習","8-3. 数学・理系基礎"),
(334,"🏢 チェーンストアのマネジメント","2026/01/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(335,"🧬 リバースストーリボード","2026/01/12","9. エンタメ・クリエイティブ","9-1. 映画・アニメ・脚本"),
(336,"🛵 Scooter Engineering and Mechanics: A Comprehensive Technical Analysis","2026/01/12","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(337,"🤩 250913AI","2025/09/13","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(338,"🎬 GeminiやGrokやSunoやCapCutといった最新のAIツールを組み合わせ高品質な2分間の動画","2026/01/12","9. エンタメ・クリエイティブ","9-4. メディア・出版・映像制作"),
(339,"⏳ Geroscience: Redefining Aging as a Treatable Medical Condition","2026/01/12","7. 健康・医療・ウェルネス","7-2. 老化・セノリティクス"),
(340,"🧟‍♂️ The Dawn of Senolytics: Eradicating Zombie Cells for Longevity","2026/01/12","7. 健康・医療・ウェルネス","7-2. 老化・セノリティクス"),
(341,"🧬 第一三共cm","2026/01/10","9. エンタメ・クリエイティブ","9-3. 広告クリエイター・デザイナー"),
(342,"🧻 BtoBとBtoCの断絶：大王製紙にみるブランド戦略の転換","2026/01/04","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(343,"📦 レンゴー株式会社とテトラパック・グループ","2026/01/04","6. 科学・工学・エネルギー","6-4. パッケージング・物流"),
(344,"⚡ 日本のエネルギー需給の現状と将来の展望","2026/01/03","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(345,"🍜 宝産業株式会社","2026/01/03","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(346,"💡 トーマス・エジソンによる白熱電球の実用化プロセス","2026/01/03","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(347,"🛵 バイク冬の寒さ対策","2026/01/03","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(348,"🛵 スクーター事務所化","2026/01/03","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(349,"🏨 星野リゾート：運営の達人が築く包括的競争戦略分析","2026/01/03","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(350,"🍽️ 2025-2026 日本飲食業界：構造変革と未来予測レポート","2026/01/03","4. 飲食・食品産業","4-1. 外食チェーン分析"),
(351,"🔬 ASML","2026/01/03","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(352,"🍣 Large Tuna Dissection: Anatomy Technique and Value Structure","2026/01/02","4. 飲食・食品産業","4-2. 食品製造・食文化"),
(353,"♨️ The Yutonami Model: Revitalizing Japan's Sento Through Business Succession","2026/01/02","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(354,"📟 TSMC: Secrets of the World-Shifting Giant","2026/01/02","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(355,"🌊 Toffler's Third Wave: A Roadmap for Digital Civilization","2026/01/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(356,"🛍️ Retail Revolution: The Rise of Specialists and Strategy in Japan","2026/01/02","2. ビジネス・経営戦略","2-3. 企業分析（消費財・サービス）"),
(357,"📖 Mastering Natural Chinese Conversation Strategies","2026/01/02","8. 教育・学習","8-1. 語学学習"),
(358,"✏️ The Universal Success of Kumon Education Methods","2026/01/02","8. 教育・学習","8-2. 教育理論・AI教育"),
(359,"🐒 【本】サピエンス全史","2025/12/30","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(360,"🏠 A Guide to Wooden Residential Construction Costs","2026/01/02","3. 建築・空間デザイン","3-2. 建築技術・工法・法規制"),
(361,"📖 Mastering Chinese Grammar for Conversational Fluency","2026/01/02","8. 教育・学習","8-1. 語学学習"),
(362,"🚀 The 45-Second Presentation: Principles of Network Organization Construction","2026/01/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(363,"🇨🇳 Learning Chinese with the Three Kingdoms: Beginner Edition","2026/01/02","8. 教育・学習","8-1. 語学学習"),
(364,"📔 Untitled notebook","2026/01/02","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(365,"🛤️ Global Railway Gauge Standards and Variations","2026/01/02","6. 科学・工学・エネルギー","6-3. 乗り物・モビリティ"),
(366,"⚖️ 波頭亮","2025/11/07","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(367,"🎯 Competition-Free Management: The Niche Strategy","2025/12/01","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(368,"🌀 【idea】AI駆動型階層化学習：NotebookLMによる教育モデルの再構築","2025/12/29","8. 教育・学習","8-2. 教育理論・AI教育"),
(369,"🧬 Senox and the Science of Senolytics","2025/12/29","7. 健康・医療・ウェルネス","7-2. 老化・セノリティクス"),
(370,"⚖️ AI Transformation: Navigating Organizational Barriers and Incentives","2025/12/22","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(371,"⚖️ AI Implementation Strategy: Lessons from Mercari and Hayakawa Gomi","2025/12/22","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(372,"🗡️ 三国志","2025/12/07","9. エンタメ・クリエイティブ","9-2. ゲーム・ボードゲーム"),
(373,"🤖 ものづくり太郎","2025/11/26","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(374,"💰 Google's AI Supremacy: Scaling and Ecosystem Advantage","2025/12/07","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(375,"🛠️ Makers","2025/12/02","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(376,"🚀 AI","2025/12/04","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(377,"👑 教育論","2025/12/03","8. 教育・学習","8-2. 教育理論・AI教育"),
(378,"🌎 外国人問題","2025/12/02","10. ライフスタイル・地方創生","10-2. 住まい・暮らし"),
(379,"🎯 電子部品20%利益率を実現するニッチ戦略","2025/12/01","2. ビジネス・経営戦略","2-1. 経営理論・フレームワーク"),
(380,"⌛ Minimal Fab","2025/11/26","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(381,"💡 Minimal Fab: Driving New Advanced Packaging and Energy Efficiency","2025/11/26","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(382,"🪄 ミニマルファブ","2025/11/26","1. テクノロジー・AI","1-3. 半導体・ハードウェア"),
(383,"🤖 ロボットとAI","2025/10/20","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
(384,"🔋 固体電池が変えるEVと日本の未来","2025/10/20","6. 科学・工学・エネルギー","6-2. エネルギー・電力"),
(385,"🎓 Stanford CS230 Introduction to Deep Learning","2025/10/09","8. 教育・学習","8-2. 教育理論・AI教育"),
(386,"🤖 AI Task Automation for Small Businesses","2025/09/13","1. テクノロジー・AI","1-1. AI基盤・エージェント・LLM"),
]

# Color fills for categories
cat_colors = {
    "1. テクノロジー・AI": "D6E4F0",
    "2. ビジネス・経営戦略": "E2EFDA",
    "3. 建築・空間デザイン": "FCE4D6",
    "4. 飲食・食品産業": "FFF2CC",
    "5. 地政学・国際関係": "F2DCDB",
    "6. 科学・工学・エネルギー": "E4DFEC",
    "7. 健康・医療・ウェルネス": "D9F2E6",
    "8. 教育・学習": "DEEBF7",
    "9. エンタメ・クリエイティブ": "FDE9D9",
    "10. ライフスタイル・地方創生": "F2F2F2",
}

for i, row in enumerate(data, 2):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=c, value=val)
        cell.font = data_font
        cell.border = thin_border
        if c == 1:
            cell.alignment = Alignment(horizontal="center")
        cat = row[3]
        if cat in cat_colors:
            cell.fill = PatternFill("solid", fgColor=cat_colors[cat])

# --- Sheet 2: ロジックツリー ---
ws2 = wb.create_sheet("ロジックツリー")

tree_headers = ["大分類", "中分類", "件数（概算）"]
tree_widths = [30, 35, 15]

for c, (h, w) in enumerate(zip(tree_headers, tree_widths), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(c)].width = w

ws2.freeze_panes = "A2"

# Count from actual data
from collections import Counter
cat_counter = Counter()
subcat_counter = Counter()
for row in data:
    cat_counter[row[3]] += 1
    subcat_counter[(row[3], row[4])] += 1

tree_data = [
    ("1. テクノロジー・AI", "1-1. AI基盤・エージェント・LLM"),
    ("1. テクノロジー・AI", "1-2. Palantir・データ基盤"),
    ("1. テクノロジー・AI", "1-3. 半導体・ハードウェア"),
    ("1. テクノロジー・AI", "1-4. ソフトウェア開発・Vibe Coding"),
    ("1. テクノロジー・AI", "1-5. デバイス・ディスプレイ"),
    ("2. ビジネス・経営戦略", "2-1. 経営理論・フレームワーク"),
    ("2. ビジネス・経営戦略", "2-2. 企業分析（テック・金融）"),
    ("2. ビジネス・経営戦略", "2-3. 企業分析（消費財・サービス）"),
    ("2. ビジネス・経営戦略", "2-4. 不動産・投資"),
    ("2. ビジネス・経営戦略", "2-5. 広告・マーケティング"),
    ("3. 建築・空間デザイン", "3-1. 体験設計・イマーシブデザイン"),
    ("3. 建築・空間デザイン", "3-2. 建築技術・工法・法規制"),
    ("3. 建築・空間デザイン", "3-3. 工務店×AI活用"),
    ("3. 建築・空間デザイン", "3-4. 都市開発・インフラ"),
    ("4. 飲食・食品産業", "4-1. 外食チェーン分析"),
    ("4. 飲食・食品産業", "4-2. 食品製造・食文化"),
    ("4. 飲食・食品産業", "4-3. 冷凍・冷蔵・物流"),
    ("5. 地政学・国際関係", "5-1. 中東・軍事・インテリジェンス"),
    ("5. 地政学・国際関係", "5-2. グローバル経済・地政学"),
    ("6. 科学・工学・エネルギー", "6-1. 物理・化学・素材"),
    ("6. 科学・工学・エネルギー", "6-2. エネルギー・電力"),
    ("6. 科学・工学・エネルギー", "6-3. 乗り物・モビリティ"),
    ("6. 科学・工学・エネルギー", "6-4. パッケージング・物流"),
    ("7. 健康・医療・ウェルネス", "7-1. 糖質制限・ダイエット"),
    ("7. 健康・医療・ウェルネス", "7-2. 老化・セノリティクス"),
    ("7. 健康・医療・ウェルネス", "7-3. 運動・脳科学"),
    ("8. 教育・学習", "8-1. 語学学習"),
    ("8. 教育・学習", "8-2. 教育理論・AI教育"),
    ("8. 教育・学習", "8-3. 数学・理系基礎"),
    ("9. エンタメ・クリエイティブ", "9-1. 映画・アニメ・脚本"),
    ("9. エンタメ・クリエイティブ", "9-2. ゲーム・ボードゲーム"),
    ("9. エンタメ・クリエイティブ", "9-3. 広告クリエイター・デザイナー"),
    ("9. エンタメ・クリエイティブ", "9-4. メディア・出版・映像制作"),
    ("10. ライフスタイル・地方創生", "10-1. 二拠点生活・地方創生"),
    ("10. ライフスタイル・地方創生", "10-2. 住まい・暮らし"),
]

for i, (cat, subcat) in enumerate(tree_data, 2):
    count = subcat_counter.get((cat, subcat), 0)
    ws2.cell(row=i, column=1, value=cat).font = data_font
    ws2.cell(row=i, column=2, value=subcat).font = data_font
    ws2.cell(row=i, column=3, value=count).font = data_font
    ws2.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    for c in range(1, 4):
        ws2.cell(row=i, column=c).border = thin_border
        if cat in cat_colors:
            ws2.cell(row=i, column=c).fill = PatternFill("solid", fgColor=cat_colors[cat])

# Summary row
summary_row = len(tree_data) + 3
ws2.cell(row=summary_row, column=1, value="合計").font = Font(name="Arial", bold=True, size=11)
ws2.cell(row=summary_row, column=3, value=f"=SUM(C2:C{len(tree_data)+1})").font = Font(name="Arial", bold=True, size=11)
ws2.cell(row=summary_row, column=3).alignment = Alignment(horizontal="center")

# Category subtotals
subtotal_row = summary_row + 2
ws2.cell(row=subtotal_row, column=1, value="大分類別集計").font = Font(name="Arial", bold=True, size=11, color="2F5496")
ws2.cell(row=subtotal_row, column=1).fill = PatternFill("solid", fgColor="D6E4F0")

cats_ordered = list(dict.fromkeys([t[0] for t in tree_data]))
for j, cat in enumerate(cats_ordered):
    r = subtotal_row + 1 + j
    ws2.cell(row=r, column=1, value=cat).font = data_font
    ws2.cell(row=r, column=3, value=cat_counter[cat]).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    for c in range(1, 4):
        ws2.cell(row=r, column=c).border = thin_border
        if cat in cat_colors:
            ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=cat_colors[cat])

output = r"C:\Users\admin\Desktop\NotebookLM_一覧.xlsx"
wb.save(output)
print(f"Saved: {output}")
