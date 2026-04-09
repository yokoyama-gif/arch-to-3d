import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')
ws = wb['判定表']

# Find a row with ○ value to get green fill
print('=== Colors of all verdict symbols in existing rows ===')
seen = {}
for r in range(6, 46):
    for c in range(10, 38):
        v = ws.cell(r, c).value
        fc = ws.cell(r, c).fill.fgColor.rgb if ws.cell(r, c).fill.fgColor else None
        fnt = ws.cell(r, c).font.color.rgb if ws.cell(r, c).font.color else None
        if v not in seen:
            seen[v] = (fc, fnt, r, c)
for k, v in seen.items():
    print(f'  {k!r}: fill={v[0]}, font={v[1]}, sample=R{v[2]}C{v[3]}')

# Check the newly added rows with ○
print('\n=== R47 vs R48 mismatch check ===')
for r in [47, 48, 56, 62]:
    print(f'Row {r}:')
    for c in [10, 11, 12, 13, 29, 31, 32, 37]:
        cell = ws.cell(r, c)
        fc = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        print(f'  C{c}: val={cell.value!r} fill={fc}')
