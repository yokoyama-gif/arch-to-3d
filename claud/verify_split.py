import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')

for name in wb.sheetnames:
    ws = wb[name]
    print(f'=== {name} ({ws.max_row}x{ws.max_column}) ===')

# Verify fill colors match values on the 18 sheet
ws = wb['18']
print('\n=== 18 sheet verdict fill check ===')
mismatches = 0
for r in range(6, ws.max_row+1):
    for c in range(10, ws.max_column+1):
        cell = ws.cell(r, c)
        v = cell.value
        fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        expected = {'○':'00C6EFCE','×':'00FFC7CE','要確認':'00FFEB9C'}.get(v)
        if expected and fill != expected:
            mismatches += 1
            if mismatches < 5:
                print(f'  MISMATCH R{r}C{c}: val={v!r} fill={fill} expected={expected}')
print(f'Total mismatches on 18 sheet: {mismatches}')

# Same for 14
ws = wb['14']
mismatches = 0
for r in range(6, ws.max_row+1):
    for c in range(10, ws.max_column+1):
        cell = ws.cell(r, c)
        v = cell.value
        fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        expected = {'○':'00C6EFCE','×':'00FFC7CE','要確認':'00FFEB9C'}.get(v)
        if expected and fill != expected:
            mismatches += 1
print(f'Total mismatches on 14 sheet: {mismatches}')

# And PC16
ws = wb['PC16']
mismatches = 0
for r in range(6, ws.max_row+1):
    for c in range(10, ws.max_column+1):
        cell = ws.cell(r, c)
        v = cell.value
        fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        expected = {'○':'00C6EFCE','×':'00FFC7CE','要確認':'00FFEB9C'}.get(v)
        if expected and fill != expected:
            mismatches += 1
print(f'Total mismatches on PC16 sheet: {mismatches}')
