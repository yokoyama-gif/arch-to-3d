"""Add KMEW セラディール18mm and セラディール14mm data to kyoto_siding_checker.xlsx"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from copy import copy
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'

# (series, pattern, hinban, color_name, munsell_str, hue_num_str, hue_letter, V, C)
# C is None for neutral
DATA_18 = [
    # セラディール・親水パワーコート18 - シェードラップ
    ('セラディール・親水パワーコート18','シェードラップ','CL57811GC','ALフォグブルー','2.1PB 4.7/1.1','2.1','PB',4.7,1.1),
    ('セラディール・親水パワーコート18','シェードラップ','CL57812GC','ALフォググリーン','0.8BG 4.6/0.2','0.8','BG',4.6,0.2),
    ('セラディール・親水パワーコート18','シェードラップ','CL5781GC','ALマックスホワイト','3.7PB 8.7/0.4','3.7','PB',8.7,0.4),
    ('セラディール・親水パワーコート18','シェードラップ','CL5782GC','ALオータムホワイト','0.7Y 7.8/0.9','0.7','Y',7.8,0.9),
    ('セラディール・親水パワーコート18','シェードラップ','CL5783GC','ALダークブラウン','0.3Y 4.4/0.5','0.3','Y',4.4,0.5),
    ('セラディール・親水パワーコート18','シェードラップ','CL5784GC','ALアッシュグリーン','5.7Y 3.8/0.3','5.7','Y',3.8,0.3),
    ('セラディール・親水パワーコート18','シェードラップ','CL5785GC','ALチャコールブラウン','7.7YR 3.3/0.4','7.7','YR',3.3,0.4),
    ('セラディール・親水パワーコート18','シェードラップ','CL5786GC','ALアトランティックブルー','5PB 4.0/1.0','5','PB',4.0,1.0),
    ('セラディール・親水パワーコート18','シェードラップ','CL5787GC','ALアイボリーブラック','7.5YR 3.1/0.1','7.5','YR',3.1,0.1),
    ('セラディール・親水パワーコート18','シェードラップ','CL5788GC','ALスローグレー','3.2Y 6.9/0.4','3.2','Y',6.9,0.4),
    ('セラディール・親水パワーコート18','シェードラップ','CL5789GC','ALダスキーグレー','7.4YR 4.2/0.1','7.4','YR',4.2,0.1),
]

DATA_14 = [
    # モダンストライプ8
    ('セラディール・親水14','モダンストライプ8','CW12511GC','MWフリントブラック','2.4YR 3.5/0.1','2.4','YR',3.5,0.1),
    ('セラディール・親水14','モダンストライプ8','CW12513GC','MWシルバーグレー','6.7Y 6.5/0.3','6.7','Y',6.5,0.3),
    ('セラディール・親水14','モダンストライプ8','CW12514GC','MWベルベットブラウン','6.4YR 3.6/0.7','6.4','YR',3.6,0.7),
    ('セラディール・親水14','モダンストライプ8','CW12515GC','MWロゼレッド','6R 4.1/2.6','6','R',4.1,2.6),
    ('セラディール・親水14','モダンストライプ8','CW12516GC','MWワサビグリーン','7.5Y 5.0/2.1','7.5','Y',5.0,2.1),
    ('セラディール・親水14','モダンストライプ8','CW12518GC','MWフォグブルー','2PB 4.7/1.1','2','PB',4.7,1.1),
    ('セラディール・親水14','モダンストライプ8','CW12519GC','MWフォググリーン','2BG 4.6/0.1','2','BG',4.6,0.1),
    ('セラディール・親水14','モダンストライプ8','CW1251GC','MWオータムホワイト','0.7Y 7.8/0.9','0.7','Y',7.8,0.9),
    ('セラディール・親水14','モダンストライプ8','CW1256GC','MWアトランティックブルー','5PB 4.0/1.0','5','PB',4.0,1.0),
    ('セラディール・親水14','モダンストライプ8','CW1257GC','MWストレートグレー','N 3.5','-','N',3.5,None),
    # 板木目14
    ('セラディール・親水14','板木目14','CW1831GC','MWボワホワイト','1Y 7.0/1.5','1','Y',7.0,1.5),
    ('セラディール・親水14','板木目14','CW1832GC','MWボワグリーン','10Y 5.0/1.0','10','Y',5.0,1.0),
    ('セラディール・親水14','板木目14','CW1833GC','MWボワブラウン','6.8YR 3.7/0.8','6.8','YR',3.7,0.8),
    ('セラディール・親水14','板木目14','CW1834GC','MWボワブラック','6.8YR 3.2/0.3','6.8','YR',3.2,0.3),
    ('セラディール・親水14','板木目14','CW1835GC','MWボワグレー','0.8Y 5.6/0.9','0.8','Y',5.6,0.9),
    ('セラディール・親水14','板木目14','CW1836GC','MWボワメープル','8.3YR 5.6/3.2','8.3','YR',5.6,3.2),
    ('セラディール・親水14','板木目14','CW1838GC','MWボワイエロー','7.7YR 5.4/4.8','7.7','YR',5.4,4.8),
    ('セラディール・親水14','板木目14','CW1839GC','MWボワライトブラウン','6.6YR 3.8/1.6','6.6','YR',3.8,1.6),
    # プラシス
    ('セラディール・親水14','プラシス','CW2121GC','MWオータムホワイト','0.7Y 7.8/0.9','0.7','Y',7.8,0.9),
    ('セラディール・親水14','プラシス','CW2122GC','MWアッシュベージュ','1.6Y 7.2/1.4','1.6','Y',7.2,1.4),
    ('セラディール・親水14','プラシス','CW2125GC','MWスティルライトブラウン','0.8Y 5.5/1.0','0.8','Y',5.5,1.0),
    ('セラディール・親水14','プラシス','CW2126GC','MWチャコールブラウン','7.7YR 3.3/0.4','7.7','YR',3.3,0.4),
    ('セラディール・親水14','プラシス','CW2127GC','MWマックスホワイト','3.7PB 8.7/0.4','3.7','PB',8.7,0.4),
    # ボーダーラップ
    ('セラディール・親水14','ボーダーラップ','CW2381GC','MWマックスホワイト','3.7PB 8.7/0.4','3.7','PB',8.7,0.4),
    ('セラディール・親水14','ボーダーラップ','CW2382GC','MWオータムホワイト','0.7Y 7.8/0.9','0.7','Y',7.8,0.9),
    ('セラディール・親水14','ボーダーラップ','CW2383GC','MWダークブラウン','9.2YR 4.2/0.5','9.2','YR',4.2,0.5),
    ('セラディール・親水14','ボーダーラップ','CW2384GC','MWアッシュグリーン','2.2Y 3.8/0.4','2.2','Y',3.8,0.4),
    ('セラディール・親水14','ボーダーラップ','CW2385GC','MWチャコールブラウン','7.7YR 3.3/0.4','7.7','YR',3.3,0.4),
    ('セラディール・親水14','ボーダーラップ','CW2386GC','MWアトランティックブルー','5PB 4.0/1.0','5','PB',4.0,1.0),
    ('セラディール・親水14','ボーダーラップ','CW2387GC','MWアイボリーブラック','7.5YR 3.1/0.1','7.5','YR',3.1,0.1),
    ('セラディール・親水14','ボーダーラップ','CW2388GC','MWフォグブルー','2PB 4.7/1.1','2','PB',4.7,1.1),
    ('セラディール・親水14','ボーダーラップ','CW2389GC','MWフォググリーン','2BG 4.6/0.1','2','BG',4.6,0.1),
    # エルデフラット14
    ('セラディール・親水14','エルデフラット14','CW2421GC','MWラントホワイト','2.1Y 8.2/0.9','2.1','Y',8.2,0.9),
    ('セラディール・親水14','エルデフラット14','CW2422GC','MWラントライトグレー','4.7Y 7.2/0.3','4.7','Y',7.2,0.3),
    ('セラディール・親水14','エルデフラット14','CW2423GC','MWラントダークグレー','1.6Y 5.4/0.6','1.6','Y',5.4,0.6),
    ('セラディール・親水14','エルデフラット14','CW2424GC','MWラントブラック','3.9P 3.4/0.1','3.9','P',3.4,0.1),
    # ベルシダー14
    ('セラディール・親水14','ベルシダー14','CW2491GC','MWメロウベージュ','8.5YR 5.7/3.4','8.5','YR',5.7,3.4),
    ('セラディール・親水14','ベルシダー14','CW2492GC','MWメロウライトブラウン','8.2YR 4.7/2.0','8.2','YR',4.7,2.0),
    ('セラディール・親水14','ベルシダー14','CW2493GC','MWメロウミドルブラウン','6.7YR 3.8/1.7','6.7','YR',3.8,1.7),
    ('セラディール・親水14','ベルシダー14','CW2494GC','MWメロウダークブラウン','7YR 3.6/0.8','7','YR',3.6,0.8),
    ('セラディール・親水14','ベルシダー14','CW2495GC','MWメロウグレー','1.3Y 5.7/0.8','1.3','Y',5.7,0.8),
    ('セラディール・親水14','ベルシダー14','CW2496GC','MWメロウグリーン','4.8Y 4.1/0.3','4.8','Y',4.1,0.3),
]

ALL_DATA = DATA_18 + DATA_14

# Judgment logic
def judge_abc(hue, V, C):
    """Returns (A,B,C) tuple of (verdict, reason). S is always '要確認'"""
    Cv = 0 if C is None else C
    def jA():
        if hue not in ('R','YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jB():
        if hue not in ('YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jC():
        if hue not in ('YR','Y','P','PB','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    return jA(), jB(), jC()

# District category mapping for columns 10..37 in 判定表
# Derived from row 5 codes already in the file
DISTRICT_CATS = ['A','A','A','B','B','B','S','S','S','B','B','B','B','B','B','B','B','B','B','C','B','C','C','A','A','A','A','C']

def build_row(no, series, pattern, hinban, color, munsell, hue, V, C):
    A, B, Cc = judge_abc(hue, V, C)
    def verdict_for(cat):
        if cat == 'A': return A[0]
        if cat == 'B': return B[0]
        if cat == 'C': return Cc[0]
        if cat == 'S': return '要確認'
    cols = [no, None, pattern, color, hinban, munsell, hue, V, (C if C is not None else '-')]
    cols += [verdict_for(c) for c in DISTRICT_CATS]
    return cols, A, B, Cc

wb = openpyxl.load_workbook(XLSX)

# --- Update 判定表 ---
ws = wb['判定表']
# Update title
ws.cell(1, 1).value = '京都市景観地区 × KMEW セラディール（親水PC16・18・14）適合判定表'

# Capture template style from row 6 (existing data row)
template_row = 6
start_row = ws.max_row + 1  # start after existing data
# find first empty row
while ws.cell(start_row-1, 1).value is None and start_row > 6:
    start_row -= 1

# Actually, just append after row 46
start_row = 47
next_no = 42

print(f'Appending to 判定表 starting at row {start_row}')

for i, rec in enumerate(ALL_DATA):
    series, pattern, hinban, color, munsell, hue_num, hue, V, C = rec
    row_idx = start_row + i
    cols, A, B, Cc = build_row(next_no + i, series, pattern, hinban, color, munsell, hue, V, C)
    for c, v in enumerate(cols, start=1):
        cell = ws.cell(row_idx, c)
        cell.value = v
        # copy style from template row
        tcell = ws.cell(template_row, c)
        if tcell.has_style:
            cell.font = copy(tcell.font)
            cell.alignment = copy(tcell.alignment)
            cell.border = copy(tcell.border)
            cell.fill = copy(tcell.fill)
            cell.number_format = tcell.number_format

# --- Update サイディング別ABC判定 ---
ws2 = wb['サイディング別ABC判定']
ws2.cell(1, 1).value = 'KMEW セラディール（親水PC16・18・14）― 色区分別 適合サマリー'

template_row2 = 5
start_row2 = 46
print(f'Appending to サイディング別ABC判定 starting at row {start_row2}')

for i, rec in enumerate(ALL_DATA):
    series, pattern, hinban, color, munsell, hue_num, hue, V, C = rec
    row_idx = start_row2 + i
    _, A, B, Cc = build_row(next_no + i, series, pattern, hinban, color, munsell, hue, V, C)
    cols2 = [next_no + i, None, pattern, color, hinban, munsell, hue, V, (C if C is not None else '-'),
             None, A[0], A[1], B[0], B[1], Cc[0], Cc[1]]
    for c, v in enumerate(cols2, start=1):
        cell = ws2.cell(row_idx, c)
        cell.value = v
        tcell = ws2.cell(template_row2, c)
        if tcell.has_style:
            cell.font = copy(tcell.font)
            cell.alignment = copy(tcell.alignment)
            cell.border = copy(tcell.border)
            cell.fill = copy(tcell.fill)
            cell.number_format = tcell.number_format

wb.save(XLSX)
print('Saved successfully.')
print(f'Added {len(ALL_DATA)} rows to each sheet.')
print(f'  - セラディール・親水パワーコート18: {len(DATA_18)} colors')
print(f'  - セラディール・親水14: {len(DATA_14)} colors')
