"""Download KMEW product images and embed in column B of each judgment sheet."""
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import urllib.request, os, io, sys, re

sys.stdout.reconfigure(encoding='utf-8')

try:
    from PIL import Image as PILImage
except ImportError:
    print('Installing Pillow...')
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', 'Pillow'])
    from PIL import Image as PILImage

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
CACHE = 'C:/Users/admin/Desktop/claud/kmew_images'
os.makedirs(CACHE, exist_ok=True)

THUMB = 72  # thumbnail px size

def try_fetch(hinban):
    """Fetch image bytes for a hinban, or None."""
    cache_path = os.path.join(CACHE, f'{hinban}.jpg')
    if os.path.exists(cache_path):
        if os.path.getsize(cache_path) > 0:
            return cache_path
        return None
    for size in [200, 72]:
        url = f'https://www.kmew.co.jp/image_kmew/gaiheki/{size}/{hinban}_{size}.jpg'
        try:
            req = urllib.request.Request(url, headers={'User-Agent':'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=15) as r:
                data = r.read()
            if len(data) > 100:
                with open(cache_path, 'wb') as f:
                    f.write(data)
                return cache_path
        except Exception:
            continue
    # Mark as failed (empty file) to avoid retry
    open(cache_path, 'wb').close()
    return None

def get_thumbnail(hinbans):
    """Try each hinban variant, return path to resized thumbnail."""
    candidates = []
    for h in hinbans:
        candidates.append(h)
        # also try without trailing A → K/U
        if h.endswith('A'):
            candidates.append(h[:-1] + 'K')
            candidates.append(h[:-1] + 'U')
    for h in candidates:
        p = try_fetch(h)
        if p:
            thumb_path = os.path.join(CACHE, f'{h}_thumb.png')
            if not os.path.exists(thumb_path):
                try:
                    img = PILImage.open(p).convert('RGB')
                    img.thumbnail((THUMB, THUMB), PILImage.LANCZOS)
                    img.save(thumb_path, 'PNG')
                except Exception as e:
                    print(f'  resize fail {h}: {e}')
                    continue
            return thumb_path
    return None

wb = openpyxl.load_workbook(XLSX)

# All judgment sheets (not ABC, not 地区基準, not 使い方)
JUDGMENT_SHEETS = [s for s in wb.sheetnames
                   if not s.startswith('ABC') and s not in ('地区基準一覧','使い方')]

print(f'Judgment sheets to process: {len(JUDGMENT_SHEETS)}')

total_added = 0
total_missing = 0

for sheet_name in JUDGMENT_SHEETS:
    ws = wb[sheet_name]
    added = 0
    missing = 0
    # Set row height + column B width for image rows
    ws.column_dimensions['B'].width = 12
    for r in range(6, ws.max_row + 1):
        hinban_cell = ws.cell(r, 5).value  # 品番 column
        if not hinban_cell:
            continue
        hinbans = [h.strip() for h in re.split(r'[/,、]', str(hinban_cell)) if h.strip()]
        thumb = get_thumbnail(hinbans)
        if thumb:
            ws.row_dimensions[r].height = 54  # ~72px
            xl_img = XLImage(thumb)
            xl_img.width = 68
            xl_img.height = 68
            xl_img.anchor = f'B{r}'
            ws.add_image(xl_img)
            added += 1
        else:
            missing += 1
    print(f'  {sheet_name}: added={added} missing={missing}')
    total_added += added
    total_missing += missing

wb.save(XLSX)
print(f'\nTotal added: {total_added}, missing: {total_missing}')
