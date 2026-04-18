"""Create an index navigation sheet + back buttons on all sheets."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
INDEX_NAME = '目次'

wb = openpyxl.load_workbook(XLSX)

# Remove existing index if any
if INDEX_NAME in wb.sheetnames:
    del wb[INDEX_NAME]

# Category mapping for grouping
def categorize(name):
    if name in ('地区基準一覧', '使い方'):
        return ('補助', 3)
    if name.startswith('ABC'):
        return ('ABC判定 (A/B/C色区分サマリー)', 2)
    return ('判定表 (京都市 景観地区別 適合判定)', 1)

groups = {}
for s in wb.sheetnames:
    cat, order = categorize(s)
    groups.setdefault((order, cat), []).append(s)

# Create index sheet at position 0
ws = wb.create_sheet(INDEX_NAME, 0)

# Styles
title_font = Font(name='游ゴシック', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
header_font = Font(name='游ゴシック', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
link_font = Font(name='游ゴシック', size=11, color='0563C1', underline='single')
cell_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center', indent=1)

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 50

# Title
ws.merge_cells('A1:D1')
ws['A1'] = '京都市景観地区 × KMEW サイディング適合判定表  ― 目次'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:D2')
ws['A2'] = 'シート名をクリックすると該当シートへジャンプします'
ws['A2'].font = Font(name='游ゴシック', size=10, italic=True, color='595959')
ws['A2'].alignment = center
ws.row_dimensions[2].height = 18

r = 4
for (order, cat), sheets in sorted(groups.items()):
    # Group header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(r, 1)
    cell.value = cat
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = left
    ws.row_dimensions[r].height = 22
    r += 1

    # Column labels
    headers = ['#', 'シート名', '行数', '備考']
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i)
        c.value = h
        c.font = Font(name='游ゴシック', size=10, bold=True)
        c.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        c.alignment = center
        c.border = cell_border
    r += 1

    # Rows
    for i, sname in enumerate(sheets, 1):
        target_ws = wb[sname]
        data_rows = max(0, target_ws.max_row - 5)

        # Get title from target sheet A1 if available
        desc = ''
        title_cell = target_ws['A1'].value
        if title_cell:
            desc = str(title_cell)
            if len(desc) > 60:
                desc = desc[:58] + '…'

        ws.cell(r, 1, i).alignment = center
        ws.cell(r, 1).border = cell_border

        name_cell = ws.cell(r, 2)
        name_cell.value = sname
        safe_name = sname.replace("'", "''")
        name_cell.hyperlink = f"#'{safe_name}'!A1"
        name_cell.font = link_font
        name_cell.alignment = left
        name_cell.border = cell_border

        cnt_cell = ws.cell(r, 3, data_rows)
        cnt_cell.alignment = center
        cnt_cell.border = cell_border

        desc_cell = ws.cell(r, 4, desc)
        desc_cell.font = Font(name='游ゴシック', size=9, color='595959')
        desc_cell.alignment = left
        desc_cell.border = cell_border

        r += 1
    r += 1  # blank row between groups

# Freeze title area
ws.freeze_panes = 'A4'

# ============================================================
# Add "← 目次に戻る" button on every other sheet (A2 merged cell area)
# ============================================================
BACK_LABEL = '← 目次に戻る'

for sname in wb.sheetnames:
    if sname == INDEX_NAME:
        continue
    ws2 = wb[sname]

    # Insert a new row at the top to host the back button (row 1 becomes row 2)
    # To avoid disturbing existing header layout, we place button in a free cell.
    # Use the last column + 1 as the button, OR insert at top. Let's try top.
    # Better: use the first available cell at a fixed top-right position using
    # unused cell like column just after data. Instead, put it as a small box
    # at top-left row 2 column that's typically empty.

    # Place at cell A1 may conflict with title. Use cell after title (e.g. J2)
    # Safest: use a merged block. Let's place at the rightmost visible area.
    # Simpler: put the link in row 2, rightmost column (max_column).
    target_col = min(ws2.max_column, 37)
    target_cell = ws2.cell(2, target_col)
    # If that cell has content, use cell 3 of same col. Else use it.
    if target_cell.value not in (None, ''):
        # Try cell at (row=3, col=target_col)
        target_cell = ws2.cell(3, target_col)

    target_cell.value = BACK_LABEL
    target_cell.hyperlink = f"#'{INDEX_NAME}'!A1"
    target_cell.font = Font(name='游ゴシック', size=10, bold=True,
                            color='FFFFFF', underline='single')
    target_cell.fill = PatternFill(start_color='C00000', end_color='C00000',
                                   fill_type='solid')
    target_cell.alignment = Alignment(horizontal='center', vertical='center')
    target_cell.border = Border(
        left=Side(style='medium', color='7F0000'),
        right=Side(style='medium', color='7F0000'),
        top=Side(style='medium', color='7F0000'),
        bottom=Side(style='medium', color='7F0000'),
    )

wb.save(XLSX)
print(f'目次シート作成完了。全 {len(wb.sheetnames)} シート')
print(f'目次: "{INDEX_NAME}" (先頭)')
print(f'各シート右上に「{BACK_LABEL}」ボタン追加')
