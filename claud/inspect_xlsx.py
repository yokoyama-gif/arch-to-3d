import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')

print('=== 判定表 header rows (1-5) ===')
ws = wb['判定表']
for r in range(1, 6):
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            print(f'  R{r}C{c}: {v!r}')

print('\n=== 判定表 sample rows 6-8 ===')
for r in range(6, 9):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    print(f'R{r}:', row)

print('\n=== 判定表 last rows ===')
for r in range(ws.max_row-3, ws.max_row+1):
    row = [ws.cell(r, c).value for c in range(1, 10)]
    print(f'R{r}:', row)
