"""Add images only to remaining sheets (金属 onwards) that errored out."""
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import urllib.request, os, sys, re

sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
CACHE = 'C:/Users/admin/Desktop/claud/kmew_img_cache'
os.makedirs(CACHE, exist_ok=True)
THUMB_PX = 60

def try_download(hinban):
    raw_path = os.path.join(CACHE, f'{hinban}_raw.jpg')
    fail_path = os.path.join(CACHE, f'{hinban}_FAIL')
    if os.path.exists(fail_path):
        return None
    if os.path.exists(raw_path) and os.path.getsize(raw_path) > 100:
        return raw_path
    for size in [200, 72]:
        url = f'https://www.kmew.co.jp/image_kmew/gaiheki/{size}/{hinban}_{size}.jpg'
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=15) as r:
                data = r.read()
            if len(data) > 500:
                with open(raw_path, 'wb') as f:
                    f.write(data)
                return raw_path
        except Exception:
            continue
    try:
        open(fail_path, 'w').close()
    except Exception:
        pass
    return None

def make_thumbnail(raw_path, hinban):
    thumb_path = os.path.join(CACHE, f'{hinban}_thumb.png')
    if os.path.exists(thumb_path) and os.path.getsize(thumb_path) > 100:
        return thumb_path
    try:
        img = PILImage.open(raw_path).convert('RGB')
        img.thumbnail((THUMB_PX, THUMB_PX), PILImage.LANCZOS)
        img.save(thumb_path, 'PNG')
        return thumb_path
    except Exception:
        return None

def get_image_for_hinbans(hinban_str):
    cleaned = re.sub(r'\s+', ' ', str(hinban_str)).strip()
    parts = [h.strip() for h in re.split(r'[/,、 ]', cleaned) if h.strip() and len(h.strip()) > 2]
    candidates = []
    for h in parts:
        candidates.append(h)
        if h.endswith('A'):
            candidates.append(h[:-1] + 'K')
            candidates.append(h[:-1] + 'U')
    for h in candidates:
        raw = try_download(h)
        if raw:
            thumb = make_thumbnail(raw, h)
            if thumb:
                return thumb
    return None

wb = openpyxl.load_workbook(XLSX)

# Only remaining sheets that didn't get images
REMAINING = ['金属シンプル', '金属デザイン', '金属シンプルH', 'LAPWALL']

total_added = 0
total_missing = 0

for sheet_name in REMAINING:
    if sheet_name not in wb.sheetnames:
        print(f'  SKIP {sheet_name} (not found)')
        continue
    ws = wb[sheet_name]
    added = 0; missing = 0; missing_list = []
    ws.column_dimensions['B'].width = 10

    for r in range(6, ws.max_row + 1):
        hinban_val = ws.cell(r, 5).value
        if not hinban_val:
            continue
        thumb = get_image_for_hinbans(str(hinban_val))
        if thumb:
            ws.row_dimensions[r].height = 48
            xl_img = XLImage(thumb)
            xl_img.width = 56; xl_img.height = 56
            xl_img.anchor = f'B{r}'
            ws.add_image(xl_img)
            added += 1
        else:
            missing += 1
            if len(missing_list) < 3:
                missing_list.append(re.sub(r'\s+', '', str(hinban_val))[:20])
    suffix = f'  (例: {", ".join(missing_list)})' if missing_list else ''
    print(f'  {sheet_name}: +{added} img, -{missing} miss{suffix}')
    total_added += added
    total_missing += missing

print('Saving...')
wb.save(XLSX)
print(f'Done! +{total_added} images, -{total_missing} missing')
