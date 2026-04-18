"""Add all remaining KMEW series as sheets to kyoto_siding_checker.xlsx."""
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
import json, re, sys
from collections import OrderedDict

sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
DATA = json.load(open('C:/Users/admin/Desktop/claud/kmew_data_more.json', encoding='utf-8'))

VERDICT_STYLES = {
    '○': ('C6EFCE', '006100'),
    '×': ('FFC7CE', '9C0006'),
    '要確認': ('FFEB9C', '9C6500'),
}
DISTRICT_CATS = ['A','A','A','B','B','B','S','S','S','B','B','B','B','B','B','B','B','B','B','C','B','C','C','A','A','A','A','C']
HUE_RE = re.compile(r'([\d.]+)?\s*(R|YR|Y|GY|G|BG|B|PB|P|RP|N)$')

def parse_hue(raw):
    s = re.sub(r'\s+', '', raw.replace('マンセル値',''))
    if s == 'N': return ('-', 'N')
    m = HUE_RE.match(s)
    if m: return (m.group(1) or '-', m.group(2))
    return ('?', '?')

def to_float(x):
    try: return float(x)
    except: return None

def judge_abc(hue, V, C):
    Cv = 0 if C is None else C
    def jA():
        if hue not in ('R','YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jB():
        if hue not in ('YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jC():
        if hue not in ('YR','Y','P','PB','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    return jA(), jB(), jC()

def apply_verdict_style(cell):
    v = cell.value
    if v in VERDICT_STYLES:
        fh, ch = VERDICT_STYLES[v]
        cell.fill = PatternFill(start_color=fh, end_color=fh, fill_type='solid')
        e = cell.font
        cell.font = Font(name=e.name, size=e.size, bold=e.bold, italic=e.italic, color=ch)

def dedupe(rows):
    grouped = OrderedDict()
    for r in rows:
        hue_raw, V_raw, C_raw, color, series, pattern, hinban = r
        num, letter = parse_hue(hue_raw)
        V = to_float(V_raw)
        C = to_float(C_raw) if C_raw != '-' else None
        key = (pattern, color, num, letter, V, C)
        if key not in grouped:
            grouped[key] = {'pattern': pattern, 'color': color, 'hue_num': num,
                           'hue': letter, 'V': V, 'C': C, 'hinbans': []}
        grouped[key]['hinbans'].append(hinban)
    return list(grouped.values())

def munsell_str(num, hue, V, C):
    if hue == 'N': return f'N {V}' + (f'/{C}' if C is not None else '')
    return f'{num}{hue} {V}/{C if C is not None else "-"}'

wb = openpyxl.load_workbook(XLSX)
template_sheet = wb['PC16']
max_col = template_sheet.max_column
abc_template = wb['ABC判定_PC16']
abc_max_col = abc_template.max_column

# (sheet_name, abc_name, data_key, title)
NEW_SHEETS = [
    ('レジェール', 'ABC_レジェール', '次世代外装パネル\u3000レジェール', '次世代外装パネル レジェール'),
    ('FLAT', 'ABC_FLAT', 'FLAT DESIGN PANEL', 'FLAT DESIGN PANEL'),
    ('フィルN光セラ16トピア', 'ABC_フィルN光セラ16トピア', 'フィルテクトN・光セラ16 セラトピア', 'フィルテクトN・光セラ16 セラトピア'),
    ('エクセ光セラ15', 'ABC_エクセ光セラ15', 'エクセレージ・光セラ15', 'エクセレージ・光セラ15'),
    ('フィルN光セラ16', 'ABC_フィルN光セラ16', 'フィルテクトN・光セラ16', 'フィルテクトN・光セラ16'),
    ('フィルE光セラ16', 'ABC_フィルE光セラ16', 'フィルテクトE・光セラ16', 'フィルテクトE・光セラ16'),
    ('セラPC16_Web', 'ABC_セラPC16_Web', 'セラディール・親水パワーコート16', 'セラディール・親水パワーコート16 (Web取得)'),
    ('ネオ親水16', 'ABC_ネオ親水16', 'ネオロック・親水16', 'ネオロック・親水16'),
    ('エクセ親水16', 'ABC_エクセ親水16', 'エクセレージ・親水16', 'エクセレージ・親水16'),
    ('エクセ親水15', 'ABC_エクセ親水15', 'エクセレージ・親水15', 'エクセレージ・親水15'),
    ('フィルN親水16', 'ABC_フィルN親水16', 'フィルテクトN・親水16', 'フィルテクトN・親水16'),
    ('フィルE親水16', 'ABC_フィルE親水16', 'フィルテクトE・親水16', 'フィルテクトE・親水16'),
    ('金属シンプル', 'ABC_金属シンプル', 'シンプルシリーズ（金属サイディング）', 'シンプルシリーズ (金属)'),
    ('金属デザイン', 'ABC_金属デザイン', 'デザインシリーズ（金属サイディング）', 'デザインシリーズ (金属)'),
    ('金属シンプルH', 'ABC_金属シンプルH', 'シンプルシリーズH（金属サイディング）', 'シンプルシリーズH (金属)'),
    ('LAPWALL', 'ABC_LAPWALL', 'LAP-WALL', 'LAP-WALL'),
]

def build_judgment_sheet(sheet_name, title, deduped):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for r in range(1, 6):
        for c in range(1, max_col + 1):
            src = template_sheet.cell(r, c); new = ws.cell(r, c)
            new.value = src.value
            if src.has_style:
                new.font = copy(src.font); new.fill = copy(src.fill)
                new.alignment = copy(src.alignment); new.border = copy(src.border)
                new.number_format = src.number_format
    ws.cell(1, 1).value = f'京都市景観地区 × KMEW {title} 適合判定表'
    for cl, dim in template_sheet.column_dimensions.items():
        ws.column_dimensions[cl].width = dim.width
    for rn in range(1, 6):
        if rn in template_sheet.row_dimensions:
            ws.row_dimensions[rn].height = template_sheet.row_dimensions[rn].height
    ws.freeze_panes = template_sheet.freeze_panes

    out_r = 6; no = 1
    for item in deduped:
        V = item['V']; C = item['C']; hue = item['hue']
        if V is None: continue
        munsell = munsell_str(item['hue_num'], hue, V, C)
        A, B, Cc = judge_abc(hue, V, C)
        def verd(cat):
            return {'A':A[0],'B':B[0],'C':Cc[0],'S':'要確認'}[cat]
        hinban = '/'.join(sorted(set(item['hinbans'])))
        cols = [no, None, item['pattern'], item['color'], hinban, munsell, hue, V,
                (C if C is not None else '-')]
        cols += [verd(c) for c in DISTRICT_CATS]
        for c, v in enumerate(cols, start=1):
            src = template_sheet.cell(6, c); new = ws.cell(out_r, c)
            new.value = v
            if src.has_style:
                new.font = copy(src.font); new.fill = copy(src.fill)
                new.alignment = copy(src.alignment); new.border = copy(src.border)
                new.number_format = src.number_format
            if c >= 10:
                apply_verdict_style(new)
        out_r += 1; no += 1
    return out_r - 6

def build_abc_sheet(sheet_name, title, deduped):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for r in range(1, 5):
        for c in range(1, abc_max_col + 1):
            src = abc_template.cell(r, c); new = ws.cell(r, c)
            new.value = src.value
            if src.has_style:
                new.font = copy(src.font); new.fill = copy(src.fill)
                new.alignment = copy(src.alignment); new.border = copy(src.border)
                new.number_format = src.number_format
    ws.cell(1, 1).value = f'KMEW {title} ― 色区分別 適合サマリー'
    for cl, dim in abc_template.column_dimensions.items():
        ws.column_dimensions[cl].width = dim.width

    out_r = 5; no = 1
    for item in deduped:
        V = item['V']; C = item['C']; hue = item['hue']
        if V is None: continue
        munsell = munsell_str(item['hue_num'], hue, V, C)
        A, B, Cc = judge_abc(hue, V, C)
        hinban = '/'.join(sorted(set(item['hinbans'])))
        cols = [no, None, item['pattern'], item['color'], hinban, munsell, hue, V,
                (C if C is not None else '-'),
                None, A[0], A[1], B[0], B[1], Cc[0], Cc[1]]
        for c, v in enumerate(cols, start=1):
            src = abc_template.cell(5, c); new = ws.cell(out_r, c)
            new.value = v
            if src.has_style:
                new.font = copy(src.font); new.fill = copy(src.fill)
                new.alignment = copy(src.alignment); new.border = copy(src.border)
                new.number_format = src.number_format
            if c in (11, 13, 15):
                apply_verdict_style(new)
        out_r += 1; no += 1
    return out_r - 5

added = []
for sheet_name, abc_name, data_key, title in NEW_SHEETS:
    raw = DATA.get(data_key, [])
    if not raw:
        print(f'SKIP {title} (no data)')
        continue
    deduped = dedupe(raw)
    n1 = build_judgment_sheet(sheet_name, title, deduped)
    n2 = build_abc_sheet(abc_name, title, deduped)
    added.append((sheet_name, abc_name))
    print(f'{title}: raw={len(raw)} dedup={len(deduped)} → {sheet_name}({n1}), {abc_name}({n2})')

# Reorder: all 判定表 first, then all ABC, then 基準/使い方
judgment_order = ['PC16','18','14','光セラ18','セラトピア18','ディズニー18','エクセ14','エクセ14広幅']
abc_order = ['ABC判定_PC16','ABC判定_18','ABC判定_14','ABC判定_光セラ18','ABC判定_セラトピア18',
             'ABC判定_ディズニー18','ABC判定_エクセ14','ABC判定_エクセ14広幅']
for sn, an in added:
    judgment_order.append(sn)
    abc_order.append(an)
desired = judgment_order + abc_order + ['地区基準一覧', '使い方']
present = [s for s in desired if s in wb.sheetnames]
others = [s for s in wb.sheetnames if s not in present]
wb._sheets = [wb[s] for s in present + others]

wb.save(XLSX)
print(f'\nAdded {len(added)} new series. Total sheets: {len(wb.sheetnames)}')
