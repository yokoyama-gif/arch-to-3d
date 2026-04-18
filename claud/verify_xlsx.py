import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')
ws = wb['判定表']
print(f'判定表 dims: {ws.max_row} x {ws.max_column}')
print(f'Title: {ws.cell(1,1).value}')

# Verify added rows
print('\n--- First new row (R47) ---')
row = [ws.cell(47, c).value for c in range(1, ws.max_column+1)]
print(row)

print('\n--- Last row ---')
row = [ws.cell(ws.max_row, c).value for c in range(1, ws.max_column+1)]
print(row)

# Spot-check: CW12515GC MWロゼレッド 6R 4.1/2.6 → R hue, should be A=○(R ok,V=4.1,C=2.6), B=× (R not in YRYn), C=× (R not in YRYPPBN)
print('\n--- CW12515GC MWロゼレッド check ---')
for r in range(47, ws.max_row+1):
    if ws.cell(r, 5).value == 'CW12515GC':
        print([ws.cell(r, c).value for c in range(1, ws.max_column+1)])

# Spot-check: CW1257GC N 3.5 → all × due to V<4
print('\n--- CW1257GC MWストレートグレー N 3.5 ---')
for r in range(47, ws.max_row+1):
    if ws.cell(r, 5).value == 'CW1257GC':
        print([ws.cell(r, c).value for c in range(1, ws.max_column+1)])

# Count rows by series
print('\n--- Row counts ---')
from collections import Counter
patterns = Counter()
for r in range(47, ws.max_row+1):
    patterns[ws.cell(r,3).value] += 1
for k,v in patterns.items():
    print(f'  {k}: {v}')

# Sheet 2
ws2 = wb['サイディング別ABC判定']
print(f'\nサイディング別ABC判定 dims: {ws2.max_row} x {ws2.max_column}')
print(f'Title: {ws2.cell(1,1).value}')
print('\n--- R46 (first new) ---')
print([ws2.cell(46, c).value for c in range(1, ws2.max_column+1)])
print('\n--- Last row ---')
print([ws2.cell(ws2.max_row, c).value for c in range(1, ws2.max_column+1)])
