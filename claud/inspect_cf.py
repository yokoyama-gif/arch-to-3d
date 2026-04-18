import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')
ws = wb['判定表']

print('=== Conditional formatting ===')
for rng, rules in ws.conditional_formatting._cf_rules.items():
    print(f'Range: {rng.sqref if hasattr(rng, "sqref") else rng}')
    for rule in rules:
        print(f'  type={rule.type}, operator={rule.operator}, formula={rule.formula}, text={getattr(rule, "text", None)}')
        if rule.dxf and rule.dxf.font:
            print(f'    font color: {rule.dxf.font.color.rgb if rule.dxf.font.color else None}')
        if rule.dxf and rule.dxf.fill:
            print(f'    fill: {rule.dxf.fill.bgColor.rgb if rule.dxf.fill.bgColor else None}')

# Also check fills on cells in new rows
print('\n=== Fill colors on R47 (new row) ===')
for c in range(10, 20):
    cell = ws.cell(47, c)
    print(f'  C{c}: val={cell.value!r} fill={cell.fill.fgColor.rgb if cell.fill.fgColor else None}')

print('\n=== Fill colors on R6 (existing row) ===')
for c in range(10, 20):
    cell = ws.cell(6, c)
    print(f'  C{c}: val={cell.value!r} fill={cell.fill.fgColor.rgb if cell.fill.fgColor else None}')
