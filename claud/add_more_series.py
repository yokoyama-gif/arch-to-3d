"""Add more 14mm/18mm series to kyoto_siding_checker.xlsx."""
import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
import json
import re
import sys
from collections import OrderedDict

sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
DATA = json.load(open('C:/Users/admin/Desktop/claud/kmew_data.json', encoding='utf-8'))

VERDICT_STYLES = {
    '○': ('C6EFCE', '006100'),
    '×': ('FFC7CE', '9C0006'),
    '要確認': ('FFEB9C', '9C6500'),
}

DISTRICT_CATS = ['A','A','A','B','B','B','S','S','S','B','B','B','B','B','B','B','B','B','B','C','B','C','C','A','A','A','A','C']

HUE_RE = re.compile(r'([\d.]+)?\s*(R|YR|Y|GY|G|BG|B|PB|P|RP|N)$')

def parse_hue(raw):
    """Extract hue number and letter from '\\n\\t...\\n\\t7.6YR' or 'N'."""
    s = raw.replace('マンセル値','').strip()
    # Remove whitespace around
    s = re.sub(r'\s+', '', s)
    if s == 'N' or s.endswith('N') and len(s) == 1:
        return ('-', 'N')
    m = HUE_RE.match(s)
    if m:
        num = m.group(1) or '-'
        letter = m.group(2)
        return (num, letter)
    return ('?', '?')

def to_float(x):
    try: return float(x)
    except: return None

def judge_abc(hue, V, C):
    Cv = 0 if C is None else C
    def jA():
        if hue not in ('R','YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jB():
        if hue not in ('YR','Y','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if V >= 7: return ('×', f'V{V}≥7')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    def jC():
        if hue not in ('YR','Y','P','PB','N'): return ('×', f'色相{hue}不可')
        if V < 4: return ('×', f'V{V}<4')
        if Cv > 3: return ('×', f'C{Cv}>3')
        return ('○', None)
    return jA(), jB(), jC()

def apply_verdict_style(cell):
    v = cell.value
    if v in VERDICT_STYLES:
        fh, ch = VERDICT_STYLES[v]
        cell.fill = PatternFill(start_color=fh, end_color=fh, fill_type='solid')
        e = cell.font
        cell.font = Font(name=e.name, size=e.size, bold=e.bold, italic=e.italic, color=ch)

def dedupe(rows):
    """Group by (pattern, color_name, hue_num, hue_letter, V, C)."""
    grouped = OrderedDict()
    for r in rows:
        hue_raw, V_raw, C_raw, color, series, pattern, hinban = r
        num, letter = parse_hue(hue_raw)
        V = to_float(V_raw)
        C = to_float(C_raw) if C_raw != '-' else None
        key = (pattern, color, num, letter, V, C)
        if key not in grouped:
            grouped[key] = {
                'pattern': pattern,
                'color': color,
                'hue_num': num,
                'hue': letter,
                'V': V,
                'C': C,
                'hinbans': [],
                'series': series,
            }
        grouped[key]['hinbans'].append(hinban)
    return list(grouped.values())

def munsell_str(num, hue, V, C):
    if hue == 'N':
        return f'N {V}' + (f'/{C}' if C is not None else '')
    return f'{num}{hue} {V}/{C if C is not None else "-"}'

wb = openpyxl.load_workbook(XLSX)

# Use existing PC16 sheet as template for header + styles
template_sheet = wb['PC16']
max_col = template_sheet.max_column

# Configuration for new sheets: (sheet_name_for_判定表, sheet_name_for_ABC, series_key_in_data, title)
NEW_SHEETS = [
    ('光セラ18', 'ABC判定_光セラ18', 'ネオロック・光セラ18', 'ネオロック・光セラ18'),
    ('セラトピア18', 'ABC判定_セラトピア18', 'ネオロック・光セラ18　セラトピア', 'ネオロック・光セラ18 セラトピア'),
    ('ディズニー18', 'ABC判定_ディズニー18', 'ネオロック・光セラ18　セラトピア　ディズニーシリーズ', 'ネオロック・光セラ18 セラトピア ディズニーシリーズ'),
    ('エクセ14', 'ABC判定_エクセ14', 'エクセレージ・親水14', 'エクセレージ・親水14'),
    ('エクセ14広幅', 'ABC判定_エクセ14広幅', 'エクセレージ・親水14広幅', 'エクセレージ・親水14広幅'),
]

def build_judgment_sheet(sheet_name, title, deduped):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    # Copy header rows 1-5
    for r in range(1, 6):
        for c in range(1, max_col + 1):
            src_cell = template_sheet.cell(r, c)
            new_cell = ws.cell(r, c)
            new_cell.value = src_cell.value
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
    ws.cell(1, 1).value = f'京都市景観地区 × KMEW {title} 適合判定表'
    # Copy column widths
    for col_letter, dim in template_sheet.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    for rn in range(1, 6):
        if rn in template_sheet.row_dimensions:
            ws.row_dimensions[rn].height = template_sheet.row_dimensions[rn].height
    ws.freeze_panes = template_sheet.freeze_panes

    # Template style from PC16 row 6
    data_template_row = 6
    out_r = 6
    no = 1
    for item in deduped:
        V = item['V']
        C = item['C']
        hue = item['hue']
        if V is None:
            continue
        munsell = munsell_str(item['hue_num'], hue, V, C)
        A, B, Cc = judge_abc(hue, V, C)
        def verdict_for(cat):
            if cat == 'A': return A[0]
            if cat == 'B': return B[0]
            if cat == 'C': return Cc[0]
            if cat == 'S': return '要確認'
        hinban_disp = '/'.join(sorted(set(item['hinbans'])))
        cols = [no, None, item['pattern'], item['color'], hinban_disp, munsell, hue, V, (C if C is not None else '-')]
        cols += [verdict_for(c) for c in DISTRICT_CATS]
        for c, v in enumerate(cols, start=1):
            src_cell = template_sheet.cell(data_template_row, c)
            new_cell = ws.cell(out_r, c)
            new_cell.value = v
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
            if c >= 10:
                apply_verdict_style(new_cell)
        out_r += 1
        no += 1
    return out_r - 6

abc_template = wb['ABC判定_PC16']
abc_max_col = abc_template.max_column

def build_abc_sheet(sheet_name, title, deduped):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for r in range(1, 5):
        for c in range(1, abc_max_col + 1):
            src_cell = abc_template.cell(r, c)
            new_cell = ws.cell(r, c)
            new_cell.value = src_cell.value
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
    ws.cell(1, 1).value = f'KMEW {title} ― 色区分別 適合サマリー'
    for col_letter, dim in abc_template.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    data_template_row = 5
    out_r = 5
    no = 1
    for item in deduped:
        V = item['V']; C = item['C']; hue = item['hue']
        if V is None: continue
        munsell = munsell_str(item['hue_num'], hue, V, C)
        A, B, Cc = judge_abc(hue, V, C)
        hinban_disp = '/'.join(sorted(set(item['hinbans'])))
        cols = [no, None, item['pattern'], item['color'], hinban_disp, munsell, hue, V,
                (C if C is not None else '-'),
                None, A[0], A[1], B[0], B[1], Cc[0], Cc[1]]
        for c, v in enumerate(cols, start=1):
            src_cell = abc_template.cell(data_template_row, c)
            new_cell = ws.cell(out_r, c)
            new_cell.value = v
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
            if c in (11, 13, 15):
                apply_verdict_style(new_cell)
        out_r += 1
        no += 1
    return out_r - 5

for sheet_name, abc_name, data_key, title in NEW_SHEETS:
    raw = DATA[data_key]
    deduped = dedupe(raw)
    n1 = build_judgment_sheet(sheet_name, title, deduped)
    n2 = build_abc_sheet(abc_name, title, deduped)
    print(f'{title}: raw={len(raw)} → dedup={len(deduped)} → 判定表{n1}行, ABC{n2}行')

# Reorder sheets
desired_order = [
    'PC16', '18', '14',
    '光セラ18', 'セラトピア18', 'ディズニー18',
    'エクセ14', 'エクセ14広幅',
    'ABC判定_PC16', 'ABC判定_18', 'ABC判定_14',
    'ABC判定_光セラ18', 'ABC判定_セラトピア18', 'ABC判定_ディズニー18',
    'ABC判定_エクセ14', 'ABC判定_エクセ14広幅',
    '地区基準一覧', '使い方'
]
present = [s for s in desired_order if s in wb.sheetnames]
others = [s for s in wb.sheetnames if s not in present]
wb._sheets = [wb[s] for s in present + others]

wb.save(XLSX)
print('\nSaved. Final sheets:', wb.sheetnames)
