"""Set print orientation to portrait for all sheets."""
import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
import sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'
wb = openpyxl.load_workbook(XLSX)

for name in wb.sheetnames:
    ws = wb[name]
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # Fit to width for wide judgment sheets
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                  header=0.3, footer=0.3)
    print(f'  {name}: portrait + fit-to-width A4')

wb.save(XLSX)
print(f'\nDone. {len(wb.sheetnames)} sheets set to portrait.')
