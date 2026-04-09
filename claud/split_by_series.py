"""Split 判定表 and サイディング別ABC判定 by series into separate sheets,
and fix cell fill colors to match each cell's verdict."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
import sys

sys.stdout.reconfigure(encoding='utf-8')

XLSX = 'C:/Users/admin/Downloads/kyoto_siding_checker.xlsx'

# Verdict → (fill, font) styles
VERDICT_STYLES = {
    '○': ('C6EFCE', '006100'),
    '×': ('FFC7CE', '9C0006'),
    '要確認': ('FFEB9C', '9C6500'),
}

def apply_verdict_style(cell):
    v = cell.value
    if v in VERDICT_STYLES:
        fill_hex, font_hex = VERDICT_STYLES[v]
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
        # preserve existing font but override color
        existing = cell.font
        cell.font = Font(
            name=existing.name, size=existing.size, bold=existing.bold,
            italic=existing.italic, color=font_hex
        )

# Series identification: look at 柄名 to classify
PC16_PATTERNS = {'デュプレ','トワールストーン16','ラティーナ','シュランク16','フリーゼ',
                 'クローフ16','グリッドスクエア','デルガド','板木目16','グランシー','エルデフラット16'}
S18_PATTERNS = {'シェードラップ'}
S14_PATTERNS = {'モダンストライプ8','板木目14','プラシス','ボーダーラップ','エルデフラット14','ベルシダー14'}

SERIES_CONFIG = [
    ('PC16', 'セラディール親水PC16', PC16_PATTERNS),
    ('18', 'セラディール・親水パワーコート18', S18_PATTERNS),
    ('14', 'セラディール・親水14', S14_PATTERNS),
]

wb = openpyxl.load_workbook(XLSX)

# =========================================================================
# Split 判定表
# =========================================================================
src = wb['判定表']
max_col = src.max_column
max_row = src.max_row

# Collect data rows by pattern
all_rows = []  # list of (row_values_tuple, template_for_style)
for r in range(6, max_row + 1):
    vals = [src.cell(r, c).value for c in range(1, max_col + 1)]
    if vals[2] is None:
        continue
    all_rows.append(vals)

def make_series_sheet(sheet_name, series_title, patterns, base_src):
    # Remove existing sheet if present
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Copy rows 1-5 (header) from src
    for r in range(1, 6):
        for c in range(1, max_col + 1):
            src_cell = base_src.cell(r, c)
            new_cell = ws.cell(r, c)
            new_cell.value = src_cell.value
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
    # Update title
    ws.cell(1, 1).value = f'京都市景観地区 × KMEW {series_title} 適合判定表'

    # Copy column widths
    for col_letter, dim in base_src.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width
    # Copy row heights for header rows
    for rn in range(1, 6):
        if rn in base_src.row_dimensions:
            ws.row_dimensions[rn].height = base_src.row_dimensions[rn].height

    # Write data rows matching patterns
    out_r = 6
    no = 1
    for vals in all_rows:
        if vals[2] not in patterns:
            continue
        # Use a template data row from base_src (row 6) for styling of non-verdict cols
        for c in range(1, max_col + 1):
            template_cell = base_src.cell(6, c)
            new_cell = ws.cell(out_r, c)
            if c == 1:
                new_cell.value = no
            else:
                new_cell.value = vals[c - 1]
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.fill = copy(template_cell.fill)
                new_cell.alignment = copy(template_cell.alignment)
                new_cell.border = copy(template_cell.border)
                new_cell.number_format = template_cell.number_format
            # Override fill/font for verdict cells based on actual value
            if c >= 10:
                apply_verdict_style(new_cell)
        out_r += 1
        no += 1

    # Freeze panes to match source if any
    ws.freeze_panes = base_src.freeze_panes
    return out_r - 6

counts = {}
for sheet_name, title, patterns in SERIES_CONFIG:
    n = make_series_sheet(sheet_name, title, patterns, src)
    counts[sheet_name] = n
    print(f'判定表→{sheet_name}: {n} rows')

# Remove the original merged 判定表
del wb['判定表']

# Reorder: put new sheets first
desired_order = ['PC16', '18', '14', 'サイディング別ABC判定', '地区基準一覧', '使い方']

# =========================================================================
# Split サイディング別ABC判定
# =========================================================================
src2 = wb['サイディング別ABC判定']
max_col2 = src2.max_column
max_row2 = src2.max_row

all_rows2 = []
for r in range(5, max_row2 + 1):
    vals = [src2.cell(r, c).value for c in range(1, max_col2 + 1)]
    if vals[2] is None:
        continue
    all_rows2.append(vals)

def make_abc_sheet(sheet_name_suffix, series_title, patterns, base_src):
    sheet_name = f'ABC判定_{sheet_name_suffix}'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Copy rows 1-4 (header)
    for r in range(1, 5):
        for c in range(1, max_col2 + 1):
            src_cell = base_src.cell(r, c)
            new_cell = ws.cell(r, c)
            new_cell.value = src_cell.value
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.border = copy(src_cell.border)
                new_cell.number_format = src_cell.number_format
    ws.cell(1, 1).value = f'KMEW {series_title} ― 色区分別 適合サマリー'

    for col_letter, dim in base_src.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    out_r = 5
    no = 1
    for vals in all_rows2:
        if vals[2] not in patterns:
            continue
        for c in range(1, max_col2 + 1):
            template_cell = base_src.cell(5, c)
            new_cell = ws.cell(out_r, c)
            if c == 1:
                new_cell.value = no
            else:
                new_cell.value = vals[c - 1]
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.fill = copy(template_cell.fill)
                new_cell.alignment = copy(template_cell.alignment)
                new_cell.border = copy(template_cell.border)
                new_cell.number_format = template_cell.number_format
            # Apply verdict styling to A/B/C judgment columns (11, 13, 15)
            if c in (11, 13, 15):
                apply_verdict_style(new_cell)
        out_r += 1
        no += 1
    return out_r - 5

for sheet_name, title, patterns in SERIES_CONFIG:
    n = make_abc_sheet(sheet_name, title, patterns, src2)
    print(f'ABC判定→ABC判定_{sheet_name}: {n} rows')

del wb['サイディング別ABC判定']

# Reorder
desired_order = ['PC16', '18', '14',
                 'ABC判定_PC16', 'ABC判定_18', 'ABC判定_14',
                 '地区基準一覧', '使い方']
present = [s for s in desired_order if s in wb.sheetnames]
others = [s for s in wb.sheetnames if s not in present]
wb._sheets = [wb[s] for s in present + others]

wb.save(XLSX)
print('\nSaved. Final sheets:', wb.sheetnames)
