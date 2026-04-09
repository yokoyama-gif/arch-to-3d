import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')

ws = wb['地区基準一覧']
print('=== 地区基準一覧 ===')
for r in range(1, ws.max_row+1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    if any(x is not None for x in row):
        print(f'R{r}:', row)

print('\n=== サイディング別ABC判定 ===')
ws2 = wb['サイディング別ABC判定']
for r in range(1, ws2.max_row+1):
    row = [ws2.cell(r, c).value for c in range(1, ws2.max_column+1)]
    if any(x is not None for x in row):
        print(f'R{r}:', row)
