import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('C:/Users/admin/Downloads/kyoto_siding_checker.xlsx')
ws = wb['判定表']

# Read all data rows
for r in range(6, ws.max_row+1):
    no = ws.cell(r, 1).value
    series = ws.cell(r, 3).value
    color = ws.cell(r, 4).value
    code = ws.cell(r, 5).value
    munsell = ws.cell(r, 6).value
    hue = ws.cell(r, 7).value
    v = ws.cell(r, 8).value
    c = ws.cell(r, 9).value
    print(f'{no}|{series}|{color}|{code}|{munsell}|{hue}|{v}|{c}')

# Also read ABC sheet
print('\n--- ABC sheet ---')
ws2 = wb['サイディング別ABC判定']
# headers
for c in range(10, ws2.max_column+1):
    print(f'Col{c}: {ws2.cell(4, c).value}')

print('\n--- ABC data ---')
for r in range(5, ws2.max_row+1):
    no = ws2.cell(r, 1).value
    vals = [str(ws2.cell(r, c).value)[:20] for c in range(10, ws2.max_column+1)]
    print(f'{no}|{"|".join(vals)}')
