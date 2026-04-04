from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "YouTube再生履歴"

# Header styling
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="CC0000")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["視聴日", "タイトル", "チャンネル名", "種類", "URL"]
col_widths = [14, 70, 25, 10, 50]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[chr(64 + col)].width = w

# Data
data = [
    {"date": "2026-04-04 (金)", "title": "Kesempurnaan yang Ternoda | Ustadz Ammi Nur Baits", "channel": "anb channel", "videoId": "YBUE5v7bQJQ", "type": "動画"},
    {"date": "2026-04-04 (金)", "title": "Memahami Cara Beragama | Ustadz Ammi Nur Baits", "channel": "anb channel", "videoId": "PH5fg1XhAx0", "type": "動画"},
    {"date": "2026-04-04 (金)", "title": "【シリーズプレイバック】ドジャースvsDバックス 9.24-9.26｜MLB2025シーズン", "channel": "SPOTVNOW", "videoId": "fNXwGaImzBo", "type": "動画"},
    {"date": "2026-04-04 (金)", "title": "【まさに\"GOAT\"3戦3勝で日本人投手初のMVPを受賞！山本由伸 ワールドシリーズまとめ】ドジャースvsブルージェイズ MLB2025 ワールドシリーズ 10.26-11.2", "channel": "SPOTVNOW", "videoId": "4hGopa70MSc", "type": "動画"},
    {"date": "2026-03-12", "title": "お〜いお茶 WBCチケット抽選演出‼️", "channel": "", "videoId": "", "type": "ショート"},
    {"date": "2026-03-12", "title": "【お茶点てポーズ誕生の裏話】大谷翔平の気遣い 北山亘基考案の真相は？｜ワールド・ベースボール・クラシック（WBC）", "channel": "", "videoId": "", "type": "ショート"},
    {"date": "2026-03-12", "title": "※映像あり おーいお茶を飲んだベッツの衝撃発言が話題に！#大谷翔平 #野球 #プロ野球", "channel": "", "videoId": "", "type": "ショート"},
    {"date": "2026-03-12", "title": "大谷のスポンサーが強制撤去される...お～いお茶が餌食になった日", "channel": "", "videoId": "", "type": "ショート"},
    {"date": "2026-03-12", "title": "おーいお茶をよけるゲレーロ・ソト、宣伝するタティス", "channel": "", "videoId": "", "type": "ショート"},
    {"date": "2026-03-12", "title": "大谷の「お〜いお茶」をソトが撤去！？ゲレーロJr ＆タティスJr 爆笑😂", "channel": "TJ Sports: Extra!", "videoId": "SXZezP290cM", "type": "動画"},
]

data_font = Font(name="Arial", size=10)
date_align = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(data, 2):
    url = f"https://www.youtube.com/watch?v={row['videoId']}" if row["videoId"] else ""
    values = [row["date"], row["title"], row["channel"], row["type"], url]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = data_font
        cell.border = thin_border
        if col == 1 or col == 4:
            cell.alignment = date_align

# Freeze header
ws.freeze_panes = "A2"
# Auto-filter
ws.auto_filter.ref = f"A1:E{len(data) + 1}"

output_path = "C:/Users/admin/Desktop/youtube_history.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total: {len(data)} entries")
