#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MAX_EXCEL_CELL_LENGTH = 32767
DEFAULT_RULES_PATH = Path(__file__).with_name("category_rules.json")


@dataclass
class MessageRow:
    conversation_id: str
    title: str
    turn_index: int
    role: str
    timestamp: str
    model: str
    content: str


@dataclass
class ConversationRow:
    conversation_id: str
    title: str
    category: str
    category_code: str
    score: int
    matched_keywords: str
    created_at: str
    updated_at: str
    message_count: int
    user_message_count: int
    assistant_message_count: int
    last_model: str
    preview: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert ChatGPT exported conversations into a classified Excel workbook."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to ChatGPT export ZIP, extracted export directory, or conversations.json.",
    )
    parser.add_argument(
        "--output",
        help="Output .xlsx path. Default: <input_name>_classified.xlsx",
    )
    parser.add_argument(
        "--rules",
        default=str(DEFAULT_RULES_PATH),
        help="Path to classification rule JSON.",
    )
    parser.add_argument(
        "--timezone",
        default="Asia/Tokyo",
        help="Timezone for timestamp rendering. Example: Asia/Tokyo",
    )
    parser.add_argument(
        "--preview-chars",
        type=int,
        default=180,
        help="Number of characters to keep in the summary preview column.",
    )
    parser.add_argument(
        "--skip-messages-sheet",
        action="store_true",
        help="Skip the Messages sheet when the export is very large.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    rules_path = Path(args.rules).expanduser().resolve()

    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")
    if not rules_path.exists():
        raise SystemExit(f"Rules file not found: {rules_path}")

    tz = load_timezone(args.timezone)
    rules = load_rules(rules_path)
    conversations = load_conversations(input_path)
    summary_rows, message_rows = build_rows(
        conversations=conversations,
        rules=rules,
        tz=tz,
        preview_chars=args.preview_chars,
    )

    output_path = resolve_output_path(input_path, args.output)
    write_workbook(
        output_path=output_path,
        summary_rows=summary_rows,
        message_rows=message_rows,
        skip_messages_sheet=args.skip_messages_sheet,
    )

    print(f"Excel exported: {output_path}")
    print(f"Conversations: {len(summary_rows)}")
    print(f"Messages: {len(message_rows)}")
    return 0


def load_timezone(tz_name: str) -> ZoneInfo:
    try:
        return ZoneInfo(tz_name)
    except Exception as exc:  # pragma: no cover - depends on runtime tz database
        raise SystemExit(f"Invalid timezone '{tz_name}': {exc}") from exc


def load_rules(rules_path: Path) -> list[dict[str, Any]]:
    raw = json.loads(rules_path.read_text(encoding="utf-8"))
    rules: list[dict[str, Any]] = []
    for code, payload in raw.items():
        keywords = payload.get("keywords", [])
        rules.append(
            {
                "code": code,
                "label": payload.get("label", code),
                "keywords": [normalize_text(keyword) for keyword in keywords if keyword],
            }
        )
    return rules


def resolve_output_path(input_path: Path, output_arg: str | None) -> Path:
    if output_arg:
        output_path = Path(output_arg).expanduser().resolve()
        if output_path.suffix.lower() != ".xlsx":
            raise SystemExit("Output must end with .xlsx")
        return output_path

    name = input_path.stem if input_path.is_file() else input_path.name
    return input_path.parent / f"{name}_classified.xlsx"


def load_conversations(input_path: Path) -> list[dict[str, Any]]:
    if input_path.is_dir():
        json_path = find_conversations_json(input_path)
        return json.loads(json_path.read_text(encoding="utf-8"))

    if input_path.suffix.lower() == ".json":
        return json.loads(input_path.read_text(encoding="utf-8"))

    if input_path.suffix.lower() == ".zip":
        with zipfile.ZipFile(input_path) as archive:
            candidates = [
                name for name in archive.namelist() if name.lower().endswith("conversations.json")
            ]
            if not candidates:
                sample = ", ".join(sorted(archive.namelist())[:10])
                raise SystemExit(
                    "conversations.json was not found in the ZIP export. "
                    f"Archive entries sample: {sample}"
                )
            with archive.open(candidates[0]) as handle:
                return json.load(handle)

    raise SystemExit("Input must be a ZIP export, export directory, or conversations.json")


def find_conversations_json(root: Path) -> Path:
    matches = sorted(root.rglob("conversations.json"))
    if not matches:
        raise SystemExit(f"conversations.json not found under: {root}")
    return matches[0]


def build_rows(
    conversations: list[dict[str, Any]],
    rules: list[dict[str, Any]],
    tz: ZoneInfo,
    preview_chars: int,
) -> tuple[list[ConversationRow], list[MessageRow]]:
    summary_rows: list[ConversationRow] = []
    message_rows: list[MessageRow] = []

    for conversation in conversations:
        title = clean_text(conversation.get("title") or "(no title)")
        conversation_id = str(conversation.get("id") or "")
        messages = extract_messages(conversation, title=title, conversation_id=conversation_id, tz=tz)
        if not messages:
            continue

        message_rows.extend(messages)

        user_text = "\n".join(msg.content for msg in messages if msg.role == "user")
        classification = classify_conversation(title=title, user_text=user_text, rules=rules)
        last_model = next((msg.model for msg in reversed(messages) if msg.model), "")
        preview = clean_text(user_text or messages[0].content)[:preview_chars]

        summary_rows.append(
            ConversationRow(
                conversation_id=conversation_id,
                title=title,
                category=classification["label"],
                category_code=classification["code"],
                score=classification["score"],
                matched_keywords=", ".join(classification["matched_keywords"]),
                created_at=render_timestamp(conversation.get("create_time"), tz),
                updated_at=render_timestamp(conversation.get("update_time"), tz),
                message_count=len(messages),
                user_message_count=sum(1 for msg in messages if msg.role == "user"),
                assistant_message_count=sum(1 for msg in messages if msg.role == "assistant"),
                last_model=last_model,
                preview=preview,
            )
        )

    summary_rows.sort(key=lambda row: (row.updated_at, row.created_at, row.title), reverse=True)
    return summary_rows, message_rows


def extract_messages(conversation: dict[str, Any], title: str, conversation_id: str, tz: ZoneInfo) -> list[MessageRow]:
    mapping = conversation.get("mapping") or {}
    if not isinstance(mapping, dict):
        return []

    path_pairs = main_path_pairs(mapping, conversation.get("current_node"))
    rows = build_message_rows_from_pairs(path_pairs, title=title, conversation_id=conversation_id, tz=tz)
    if rows:
        return rows

    fallback_pairs = sorted(
        mapping.items(),
        key=lambda item: (
            extract_timestamp(item[1]),
            compute_depth(item[0], mapping),
            str(item[0]),
        ),
    )
    return build_message_rows_from_pairs(
        fallback_pairs, title=title, conversation_id=conversation_id, tz=tz
    )


def main_path_pairs(
    mapping: dict[str, dict[str, Any]], current_node: Any
) -> list[tuple[str, dict[str, Any]]]:
    if not current_node:
        return []

    pairs: list[tuple[str, dict[str, Any]]] = []
    seen: set[str] = set()
    node_id = str(current_node)
    if node_id not in mapping:
        return []

    while node_id and node_id in mapping and node_id not in seen:
        seen.add(node_id)
        node = mapping[node_id]
        pairs.append((node_id, node))
        parent = node.get("parent")
        node_id = str(parent) if parent is not None else ""

    pairs.reverse()
    return pairs


def build_message_rows_from_pairs(
    pairs: Iterable[tuple[str, dict[str, Any]]],
    title: str,
    conversation_id: str,
    tz: ZoneInfo,
) -> list[MessageRow]:
    rows: list[MessageRow] = []
    turn_index = 1
    for _, node in pairs:
        message = node.get("message") or {}
        author = message.get("author") or {}
        role = str(author.get("role") or "unknown").strip().lower()
        content = extract_content_text(message.get("content"))
        if not content:
            continue

        metadata = message.get("metadata") or {}
        rows.append(
            MessageRow(
                conversation_id=conversation_id,
                title=title,
                turn_index=turn_index,
                role=role,
                timestamp=render_timestamp(message.get("create_time") or node.get("create_time"), tz),
                model=str(
                    metadata.get("model_slug")
                    or metadata.get("default_model_slug")
                    or metadata.get("model_switcher_deny")
                    or ""
                ),
                content=content,
            )
        )
        turn_index += 1
    return rows


def extract_content_text(content: Any) -> str:
    fragments: list[str] = []

    def visit(value: Any) -> None:
        if value is None:
            return
        if isinstance(value, str):
            text = clean_text(value)
            if text:
                fragments.append(text)
            return
        if isinstance(value, list):
            for item in value:
                visit(item)
            return
        if isinstance(value, dict):
            preferred_keys = (
                "parts",
                "text",
                "result",
                "content",
                "caption",
                "summary",
                "transcript",
                "title",
                "description",
            )
            for key in preferred_keys:
                if key in value:
                    visit(value[key])
            return

    visit(content)
    joined = "\n".join(fragment for fragment in fragments if fragment)
    if len(joined) > MAX_EXCEL_CELL_LENGTH:
        return joined[: MAX_EXCEL_CELL_LENGTH - 1]
    return joined


def classify_conversation(title: str, user_text: str, rules: list[dict[str, Any]]) -> dict[str, Any]:
    normalized_title = normalize_text(title)
    normalized_user_text = normalize_text(user_text)
    best = {"code": "other", "label": "その他", "score": 0, "matched_keywords": []}

    for rule in rules:
        score = 0
        matched_keywords: list[str] = []
        for keyword in rule["keywords"]:
            if keyword and keyword in normalized_title:
                score += 3
                matched_keywords.append(keyword)
            elif keyword and keyword in normalized_user_text:
                score += 1
                matched_keywords.append(keyword)

        if score > best["score"] or (
            score == best["score"] and len(matched_keywords) > len(best["matched_keywords"])
        ):
            best = {
                "code": rule["code"],
                "label": rule["label"],
                "score": score,
                "matched_keywords": matched_keywords[:10],
            }

    if best["score"] == 0 and looks_like_code_request(user_text):
        return {
            "code": "programming",
            "label": "開発・プログラミング",
            "score": 1,
            "matched_keywords": ["code-pattern"],
        }

    return best


def looks_like_code_request(text: str) -> bool:
    lowered = normalize_text(text)
    patterns = (
        "```",
        "def ",
        "class ",
        "function ",
        "select ",
        "from ",
        "console.log",
        "traceback",
        "exception",
        "stack trace",
        "syntaxerror",
    )
    return any(pattern in lowered for pattern in patterns)


def render_timestamp(raw_value: Any, tz: ZoneInfo) -> str:
    if raw_value in (None, "", 0):
        return ""

    try:
        timestamp = float(raw_value)
    except (TypeError, ValueError):
        return str(raw_value)

    dt = datetime.fromtimestamp(timestamp, tz=timezone.utc).astimezone(tz)
    return dt.isoformat(timespec="seconds")


def extract_timestamp(node: dict[str, Any]) -> float:
    message = node.get("message") or {}
    value = message.get("create_time") or node.get("create_time") or 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def compute_depth(node_id: str, mapping: dict[str, dict[str, Any]]) -> int:
    depth = 0
    seen: set[str] = set()
    current = str(node_id)
    while current and current in mapping and current not in seen:
        seen.add(current)
        parent = mapping[current].get("parent")
        current = str(parent) if parent is not None else ""
        depth += 1
    return depth


def write_workbook(
    output_path: Path,
    summary_rows: list[ConversationRow],
    message_rows: list[MessageRow],
    skip_messages_sheet: bool,
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    write_summary_sheet(summary_sheet, summary_rows)
    if not skip_messages_sheet:
        write_messages_sheet(workbook.create_sheet("Messages"), message_rows)
    write_stats_sheet(workbook.create_sheet("Stats"), summary_rows, message_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_summary_sheet(sheet: Any, rows: list[ConversationRow]) -> None:
    headers = [
        "conversation_id",
        "title",
        "category",
        "category_code",
        "score",
        "matched_keywords",
        "created_at",
        "updated_at",
        "message_count",
        "user_message_count",
        "assistant_message_count",
        "last_model",
        "preview",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row.conversation_id,
                row.title,
                row.category,
                row.category_code,
                row.score,
                row.matched_keywords,
                row.created_at,
                row.updated_at,
                row.message_count,
                row.user_message_count,
                row.assistant_message_count,
                row.last_model,
                row.preview,
            ]
        )

    style_sheet(sheet, wrap_columns={"B", "F", "M"})


def write_messages_sheet(sheet: Any, rows: list[MessageRow]) -> None:
    headers = ["conversation_id", "title", "turn_index", "role", "timestamp", "model", "content"]
    sheet.append(headers)

    for row in rows:
        sheet.append(
            [
                row.conversation_id,
                row.title,
                row.turn_index,
                row.role,
                row.timestamp,
                row.model,
                row.content,
            ]
        )

    style_sheet(sheet, wrap_columns={"B", "G"})


def write_stats_sheet(sheet: Any, summary_rows: list[ConversationRow], message_rows: list[MessageRow]) -> None:
    category_counts = Counter(row.category for row in summary_rows)
    monthly_counts = Counter((row.updated_at or "")[:7] for row in summary_rows if row.updated_at)
    model_counts = Counter(row.last_model for row in summary_rows if row.last_model)
    role_counts = Counter(row.role for row in message_rows)

    sheet["A1"] = "Category"
    sheet["B1"] = "Count"
    row_index = 2
    for category, count in category_counts.most_common():
        sheet[f"A{row_index}"] = category
        sheet[f"B{row_index}"] = count
        row_index += 1

    sheet["D1"] = "Updated Month"
    sheet["E1"] = "Count"
    row_index = 2
    for month, count in sorted(monthly_counts.items()):
        sheet[f"D{row_index}"] = month
        sheet[f"E{row_index}"] = count
        row_index += 1

    sheet["G1"] = "Model"
    sheet["H1"] = "Count"
    row_index = 2
    for model, count in model_counts.most_common():
        sheet[f"G{row_index}"] = model
        sheet[f"H{row_index}"] = count
        row_index += 1

    sheet["J1"] = "Role"
    sheet["K1"] = "Count"
    row_index = 2
    for role, count in role_counts.most_common():
        sheet[f"J{row_index}"] = role
        sheet[f"K{row_index}"] = count
        row_index += 1

    style_sheet(sheet)


def style_sheet(sheet: Any, wrap_columns: set[str] | None = None) -> None:
    wrap_columns = wrap_columns or set()
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    plain_alignment = Alignment(vertical="top", wrap_text=False)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells[:201]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        for cell in column_cells:
            cell.alignment = wrap_alignment if column_letter in wrap_columns else plain_alignment
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def normalize_text(text: str) -> str:
    normalized = clean_text(text).casefold()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def clean_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = text.strip()
    if len(text) > MAX_EXCEL_CELL_LENGTH:
        return text[: MAX_EXCEL_CELL_LENGTH - 1]
    return text


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        raise SystemExit(130)
